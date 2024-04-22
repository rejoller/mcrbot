



import openai
import aiogram
import logging
import asyncio
import csv
from datetime import datetime
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from google.oauth2 import service_account
from googleapiclient.discovery import build
from fuzzywuzzy import fuzz
from prettytable import PrettyTable
from pretty_html_table import build_table
import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
import os
import matplotlib.pyplot as plt
import numpy as np
import tempfile
import pandas as pd
import gspread
import requests
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import html
from html import escape
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

from google.oauth2 import service_account
import aiohttp
import gspread_asyncio

#import spacy
from datetime import datetime, timedelta
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from aiogram.utils.executor import start_webhook

#from aiogram.contrib.fsm_storage.redis import RedisStorage2
from aiogram.contrib.fsm_storage.memory import MemoryStorage

from bs4 import BeautifulSoup
import markdown

def remove_markdown(text):
    """
    Удаляет разметку Markdown из текста, сохраняя пробелы и переносы строк.
    """
    # Удаляем заголовки

    # Удаляем жирный текст
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
    # Удаляем курсив


    # Удаляем код

    return text


#nlp = spacy.load("ru_core_news_sm")




user_messages = {}
additional_info_storage = {}
espd_info_storage = {}
szoreg_info_storage = {}
schools_info_storage = {}
message_storage = {}
survey_data_storage = {}

districts = ["Абанский р-н", "Ачинский р-н", "Курагинский р-н"]
response_storage = {}
bot_token = ''
bot = Bot(token=bot_token)
dp = Dispatcher(bot, storage=MemoryStorage())
info_text_storage = {}




'''
WEBHOOK_HOST = 'https://rejoller.pythonanywhere.com/'
WEBHOOK_SSL_CERT = '/home/rejoller/mcrbot/YOURPRIVATE.key'
WEBHOOK_PATH = '/my_telegram_bot'
# Веб-сервер настройки
WEBAPP_HOST = '0.0.0.0'  # Привязка к всех интерфейсам
WEBAPP_PORT = 5005       # Убедитесь, что этот порт открыт и доступен
WEBHOOK_URL = f'{WEBHOOK_HOST}{WEBHOOK_PATH}'


'''

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.dispatcher.filters.state import State, StatesGroup

class Form(StatesGroup):
    waiting_for_number = State()

from aiogram.contrib.fsm_storage.redis import RedisStorage2

import redis

storage = RedisStorage2(db=0, host='redis-11158.c304.europe-west1-2.gce.cloud.redislabs.com', port=11158, password='aoOlWYyOKYFXMozbi6b24ja0A011RuEl')

storage2 = redis.Redis(db=0, host='redis-11158.c304.europe-west1-2.gce.cloud.redislabs.com', port=11158, password='aoOlWYyOKYFXMozbi6b24ja0A011RuEl')



# Удалите или закомментируйте следующую строку, так как модель не требуется для aiogram
openai.api_key = ''


SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
SERVICE_ACCOUNT_FILE = '/home/rejoller/mcrbot/credentials.json'
SERVICE_ACCOUNT_FILE = '/home/rejoller/mcrbot/credentials_Masha.json'



creds = None
creds = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# ID таблицы
#SPREADSHEET_ID_1 = '1lA6wXSOmi6nj4HDOpFdzm4_KaUQAAakNNxOyXx7p1ZQ'

#Маша -срока из гугл таблицы
SPREADSHEET_ID_1 = '1ghoLFQ6Ydbz0QRMgCfAT2_0fktJSNI4HkHIu6qKWWbU'


SPREADSHEET_ID_2 = '1qoSebEXzO9KpIZXXDD1QSA4NSuQ07byJ'

import gspread
from google.oauth2 import service_account

# Загрузка учетных данных из файла службы (Service Account)
credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# Создание экземпляра клиента gspread с аутентификацией учетной записи службы
#gc = gspread.AsyncioAuthManager(credentials)

# Асинхронная авторизация
#await gc.authorize()


# Функция для сохранения данных пользователя
def log_user_data(user_id, first_name, last_name, username, message_text):
    file_path = 'users_data.csv'
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Проверяем, существует ли файл. Если нет, создаем его с заголовками
    try:
        with open(file_path, 'x', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Timestamp', 'User ID', 'First Name', 'Last Name', 'Username', 'Message'])
    except FileExistsError:
        pass

    # Записываем данные пользователя в файл
    with open(file_path, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([current_time, user_id, first_name, last_name, username, message_text])

# Функция для сохранения данных пользователя из сообщения
async def log_user_data_from_message(message):
    user_id = message.from_user.id
    first_name = message.from_user.first_name
    last_name = message.from_user.last_name
    username = message.from_user.username
    message_text = message.text

    log_user_data(user_id, first_name, last_name, username, message_text)

def split_message(message, max_length=4096):
    if len(message) <= max_length:
        return [message]

    messages = []
    while len(message) > max_length:
        split_index = message[:max_length].rfind('\n')
        if split_index == -1:
            split_index = max_length

        messages.append(message[:split_index])
        message = message[split_index:].lstrip()

    if message:
        messages.append(message)

    return messages



@dp.message_handler(commands=['help'])
async def handle_help_command(message: types.Message):
    await log_user_data_from_message(message)
    help_text = (
        'Введи название населенного пункта или муниципального образования, чтобы получить информацию о связи. Чтобы узнать информацию о сотовой связи, выбери /2g /3g или /4g. Чтобы получить информацию о населенных пунктах без сотовой связи жми /nomobile\n\n'
        'Для получения списка ФАП из контракта с ПАО "Ростелеком" нажми /fp\n'
        'Для получения списка точек Аг.ГОиЧС из контракта с ПАО "Ростелеком" нажми /ago\n\n'
        'Чтобы узнать о подключении к ТОРКНД, введи сообщение "тор" и наименование муниципального образования. '
        'Например, "тор Енисейский".\n'
        'Если нужна статистика по всему краю, жми /knd_kraj\n\n'
        'Чтобы узнать, кто сегодня в отпуске, жми /otpusk\n\n'
        'Если остались вопросы, пиши @rejoller.')
    await message.reply(help_text)

@dp.message_handler(commands=['knd_kraj'])
async def handle_knd_kraj_command(message: types.Message):
    await log_user_data_from_message(message)
    await message.reply('Загружаю данные')
    await handle_knd_kraj_message(message)

@dp.message_handler(commands=['fp'])
async def handle_fp_command(message: types.Message):
    # Загрузите данные из файла с информацией о населенных пунктах
    user_first_name = message.from_user.first_name
    await message.reply(f'Секундочку, {user_first_name}😌')
    log_user_data_from_message(message)
    fp_data, fp_headers = load_fp_data()

    # Передайте fp_data и fp_headers в функцию handler_fp_message
    await handler_fp_message(message, fp_data, fp_headers)
    del fp_data
    del fp_headers

@dp.message_handler(commands=['ago'])
async def handle_ago_command(message: types.Message):
    # Загрузите данные из Google Sheets с информацией об АгГОиЧС
    user_first_name = message.from_user.first_name
    await message.reply(f'Секундочку, {user_first_name}😌')
    log_user_data_from_message(message)
    aggoics_data, aggoics_headers = load_aggoics_data()

    # Передайте aggoics_data и aggoics_headers в функцию handler_aggoics_message
    await handler_aggoics_message(message, aggoics_data, aggoics_headers)
    del aggoics_data
    del aggoics_headers

@dp.message_handler(commands=['otpusk'])
async def handle_otpusk_command(message: types.Message, days_ahead=14):
    # Загрузка данных из файла с информацией об отпусках
    print ('отпуск запущен')
    #await message.reply('Загружаю данные')
    await bot.send_message(message.chat.id, '🏝Загружаю️')
    await log_user_data_from_message(message)
    otpusk_data = await load_otpusk_data()


    # Получение списка сотрудников, которые сегодня в отпуске и уходят в отпуск в ближайшие 14 дней
    employees_on_vacation, employees_starting_vacation_soon = get_employees_on_vacation(otpusk_data, days_ahead)

    response = ""

    if employees_on_vacation:
        response += '*Сегодня в отпуске*😎\n\n'
        for row in employees_on_vacation:
            response += f"{row[0]}, {row[1]}\n"
            response += f"  - Дата начала отпуска: {row[3]}\n"
            response += f"  - Дата окончания отпуска: {row[4]}\n\n"

    if employees_starting_vacation_soon:
        response += f"\n*Сотрудники, уходящие в отпуск в ближайшие {days_ahead} дней*\n\n"
        for emp_row in employees_starting_vacation_soon:
            response += f"{emp_row[0]}, {emp_row[1]}\n"
            response += f"  - Дата начала отпуска: {emp_row[3]}\n"
            response += f"  - Дата окончания отпуска: {emp_row[4]}\n\n"

    if not response:
        response = "Сегодня никто не в отпуске, и никто не уходит в отпуск в ближайшие 14 дней."

    #response += f"\n\nЕсли нужен справочник, жми /employee"

    # Отправка запроса в GPT API
    #gpt_response = await send_request_to_otpusk_command(message.chat.id, response)
    #print(response)
    messages = split_message(response)

    # Отправка обработанной информации пользователю
    for msg in messages:
        #await message.reply(msg, parse_mode='Markdown')
        await bot.send_message(message.chat.id, msg, parse_mode='Markdown')

#2

async def send_request_to_otpusk_command(chat_id, response):
    global info_text_storage
    trimmed_info = info_text_storage.get(chat_id, "")[:7000]
    messages = [
        {"role": "system", "content": "Ты сообщаешь информацию о сотрудниках министерства цифрового развития Красноярского края, которые сейчас в отпуске. Пользователь тебе будет присылать список сотрудников, а ты писать красивый обзор  на этот список. Будь лаконичен не пиши пожелание каждому, сделай одно для всех в конце. ограничение 1000 символов."},
        {"role": "user", "content": "Сегодня в отпуске:\n\n"
                                   "- Хорунов Дмитрий Сергеевич, заместитель министра - начальник отдела\n"
                                   "  - Дата начала отпуска: 13.06.2023\n"
                                   "  - Дата окончания отпуска: 30.06.2023\n\n"
                                   "- 🦄 Щербакова Татьяна Александровна, консультант\n"
                                   "  - Дата начала отпуска: 13.06.2023\n"
                                   "  - Дата окончания отпуска: 26.06.2023\n\n"
                                   "Сотрудники, уходящие в отпуск в ближайшие 14 дней:\n\n"
                                   "- Кружилина Елена Владимировна, главный специалист\n"
                                   "  - Дата начала отпуска: 26.06.2023\n"
                                   "  - Дата окончания отпуска: 14.07.2023\n\n"},

        {"role": "assistant", "content": "🌴🌞 *Сегодня в отпуске:*\n\n"
                                          "- Хорунов Дмитрий Сергеевич, заместитель министра - начальник отдела\n"
                                          "  - Дата начала отпуска: 13.06.2023\n"
                                          "  - Дата окончания отпуска: 30.06.2023\n\n"
                                          "- 🦄 Щербакова Татьяна Александровна, консультант\n"
                                          "  - Дата начала отпуска: 13.06.2023\n"
                                          "  - Дата окончания отпуска: 26.06.2023\n\n\n"
                                        "Сотрудники, уходящие в отпуск в ближайшие 14 дней:\n\n"
                                          "- Кружилина Елена Владимировна, главный специалист\n"
                                          "  - Дата начала отпуска: *26.06.2023*\n"
                                          "  - Дата окончания отпуска: *14.07.2023*\n\n"
                                          "Коллеги, хорошо вам отдохнуть и возвращайтесь скорее на работу!"
                                          "Если вам нужна информация о других сотрудниках, не стесняйтесь обращаться к справочнику - нажмите /employee"},
        {"role": "user", "content": "Отлично, а теперь для этого списка сделай примерно также, но с другими пожеланиями, после каждого сотрудника обязательно пропускай строку. даты жирным шрифтом:"},
        {"role": "user", "content": response},
    ]

    response = openai.ChatCompletion.create(
        model="gpt-4-turbo-preview",
        messages=messages,
        max_tokens=2000,
        n=1,
        temperature=1,
        stream=True,
    )
    message_content = ""  # Переменная для хранения состояния сообщения с разметкой
    message_content_no_md = ""  # Переменная для хранения состояния сообщения без разметки
    chunk_counter = 0  # Счетчик количества полученных фрагментов
    message = None  # Переменная для хранения объекта сообщения

    try:
        while True:
            chunk = next(response)  # Получаем следующий фрагмент

            if chunk["object"] == "error":  # Проверяем, является ли фрагмент ошибкой
                # обработка ошибки
                break

            delta = chunk.get("choices", [{}])[0].get("delta", {})  # Получаем дельту фрагмента
            message_delta = delta.get("content")  # Извлекаем текст дельты

            if message_delta is not None and message_delta.strip():  # Проверяем, не является ли текст пустым
                message_content += message_delta  # Добавляем содержимое дельты к текущему сообщению с разметкой
                message_content_no_md += remove_markdown(message_delta)  # Удаляем разметку и добавляем содержимое дельты к текущему сообщению без разметки
                chunk_counter += 1  # Увеличиваем счетчик фрагментов

            if chunk_counter % 20 == 0 or chunk["object"] == "chat.completion":  # Если набрано 20 фрагментов или это последний фрагмент
                if message_content_no_md:  # Проверяем, что содержимое без разметки не является пустой строкой
                    if message is None:  # Если сообщение еще не создано, создаем его
                        message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                        message = await bot.send_message(chat_id=chat_id, text=message_content_no_md)
                    else:  # Если сообщение уже было создано
                        if message.text != message_content_no_md:  # Проверяем, изменился ли текст сообщения
                            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                            try:
                                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content_no_md)
                            except aiogram.exceptions.MessageNotModified:
                                pass

            await asyncio.sleep(0.05)  # Добавляем небольшую задержку между фрагментами

    except StopIteration:
        pass

    if message_content and (message is None or message.text != message_content):
        if message is None:  # Если сообщение еще не создано, создаем его
            message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            message = await bot.send_message(chat_id=chat_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
        else:  # Если сообщение уже было создано
            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            try:
                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
            except aiogram.exceptions.MessageNotModified:
                pass

    if message_content and message is not None:
        return message.text

    return None
'''

async def send_request_to_otpusk_command(chat_id, response):
    global info_text_storage
    trimmed_info = info_text_storage.get(chat_id, "")[:7000]
    messages = [
        {"role": "system", "content": "Ты просто повторяешь текст который тебе присылает пользователь. Процент правок не более 10% от общего количества текста."},

        {"role": "user", "content": response},
    ]

    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo-16k-0613",
        messages=messages,
        max_tokens=1000,
        n=1,
        temperature=0.8,
        stream=True,
    )
    message_content = ""  # Переменная для хранения состояния сообщения
    chunk_counter = 0  # Счетчик количества полученных фрагментовF
    message = None  # Переменная для хранения объекта сообщения

    try:
        while True:
            chunk = next(response)  # Получаем следующий фрагмент

            #print("Received chunk:")
            #print(chunk)

            if chunk["object"] == "error":  # Проверяем, является ли фрагмент ошибкой
                # обработка ошибки
                break

            delta = chunk.get("choices", [{}])[0].get("delta", {})  # Получаем дельту фрагментаF
            message_delta = delta.get("content")  # Извлекаем текст дельты

            if message_delta is not None and message_delta.strip():  # Проверяем, не является ли текст пустым
                message_content += message_delta  # Добавляем содержимое дельты к текущему сообщению
                chunk_counter += 1  # Увеличиваем счетчик фрагментов

            if chunk_counter % 25 == 0 or chunk["object"] == "chat.completion":  # Если набрано 20 фрагментов или это последний фрагмент
                if message_content:  # Проверяем, что содержимое не является пустой строкой
                    if message is None:  # Если сообщение еще не создано, создаем его
                        message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                        message = await bot.send_message(chat_id=chat_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
                    else:  # Если сообщение уже было создано
                        if message.text != message_content:  # Проверяем, изменился ли текст сообщения
                            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                            try:
                                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
                            except aiogram.exceptions.MessageNotModified:
                                pass

                    #print("Current message:")
                   # print(message_content)
                else:
                    print("Empty message content, skipping...")

          # await asyncio.sleep(0.05)  # Добавляем небольшую задержку между фрагментами

    except StopIteration:
        pass

    if message_content and (message is None or message.text != message_content):
        if message is None:  # Если сообщение еще не создано, создаем его
            message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            message = await bot.send_message(chat_id=chat_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
        else:  # Если сообщение уже было создано
            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            try:
                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
            except aiogram.exceptions.MessageNotModified:
                pass



    if message_content and message is not None:
        return message.text

    return None

'''


# Функция split_message_table остается без изменений

def split_message_table(headers, data, max_message_length=4096):
    table = PrettyTable()
    table.field_names = headers

    for row in data:
        if len(row) == len(headers):
            table.add_row(row)
        else:
            print(f"Skipping row with incorrect number of values: {row}")

    table_str = table.get_string()

    messages = []
    lines = table_str.split('\n')
    current_message = lines[0] + '\n' + lines[1] + '\n'

    for row in lines[2:]:
        test_message = current_message + row + '\n'

        if len(test_message) <= max_message_length:
            current_message = test_message
        else:
            messages.append(f"<pre>{current_message.strip()}</pre>")
            current_message = lines[0] + '\n' + lines[1] + '\n' + row + '\n'

    if current_message:
        messages.append(f"<pre>{current_message.strip()}</pre>")

    return messages


#3

async def get_value(row, index, default_value=''):
    try:
        return row[index]
    except IndexError:
        return default_value


def normalize_text_v2(text):
    text = text.lower().replace('ё', 'е').replace('р-н', 'район').replace('-', ' ')
    text = re.sub(r'(N|№|No)', 'N', text, flags=re.IGNORECASE)
    text = text.replace(' район', '').strip()
    return text







from google.oauth2 import service_account
import gspread_asyncio

SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
SERVICE_ACCOUNT_FILE = '/home/rejoller/mcrbot/credentials.json'

creds = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)


headers = ['Наименование', 'Население', 'Сотовая связь', 'Интернет', 'Программа', 'Таксофон', 'СЗО (узел)']


from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


def create_excel_file(headers, data):
    wb = Workbook()
    ws = wb.active

    # Шрифт и выравнивание заголовков
    header_font = Font(name='Arial', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Шрифт и выравнивание данных
    data_font = Font(name='Arial')
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Границы ячеек
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Заливка фона для заголовков
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
        cell.fill = header_fill

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_data)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    file_name = "test_file.xlsx"
    wb.save(file_name)
    #  print(f"File saved as {file_name}")
    return file_name


from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

def create_excel_file_2(data):
    # Добавляем нумерацию в первый столбец и сортируем по количеству голосов

    #sorted_data = sorted(data, key=lambda x: int(x[1]) if (len(x) > 1 and str(x[1]).isdigit()) else 0, reverse=True)

    sorted_data = sorted(data, key=lambda x: x[4] if len(x) > 4 else 0, reverse=True)


    # Отфильтровываем пустые строки
    sorted_data = [row for row in sorted_data if any(row)]

    # Добавляем нумерацию в первый столбец
    data = [[i+1] + row for i, row in enumerate(sorted_data)]

    wb = Workbook()
    ws = wb.active

    headers = ['Позиция', 'Наименование населенного пункта', 'Количество голосов', 'Время обновления', 'Комментарий Минцифры России']

    header_font = Font(name='Arial', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial')
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Изменяем цвет заливки заголовка на светло-голубой
    header_fill = PatternFill(start_color='add8e6', end_color='add8e6', fill_type='solid')

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
        cell.fill = header_fill

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_data)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file




def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            value = str(cell.value)
            length = len(value.encode('utf-8'))
            if length > max_length:
                max_length = length

        # Настройка ширины столбца
        estimated_width = max_length * 0.7  # Умножение на коэффициент для учета разных ширин символов
        worksheet.column_dimensions[column].width = estimated_width


def convert_to_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    # Настраиваем стили для заголовков
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="AED6F1",
                              end_color="AED6F1",
                              fill_type="solid")

    # Настраиваем стили для данных
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    data_fill = PatternFill(start_color="ECECEC",
                            end_color="ECECEC",
                            fill_type="solid")

    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Применяем стили
            if row_idx == 1:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                cell.fill = header_fill
            else:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                if row_idx % 2 == 0:
                    cell.fill = data_fill

    # Вызов функции для автоматической настройки ширины столбцов
    adjust_column_width(ws)

    # Добавляем автофильтр
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Закрепляем строку заголовка
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@dp.message_handler(commands=['start', 'help'])
async def send_welcome(message: types.Message):
    await message.reply("Привет! Я бот для поиска данных. Введи запрос и я постараюсь найти нужные данные.")







#4

import folium
from aiogram.types import InputFile
from folium.plugins import MarkerCluster

def load_goroda_data():
    # Загружаем данные из Google Sheets
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID_1, range='goroda2.0!A1:T1721').execute()
    rows = result.get('values', [])
    return rows



async def create_map_with_markers(rows):
    map_with_markers = folium.Map(location=[59.664482, 91.913147], zoom_start=10)

    # Создаем кластер маркеров
    marker_cluster = MarkerCluster().add_to(map_with_markers)

    for row in rows:
        if row[7] and row[8]:  # проверяем, есть ли широта и долгота
            folium.Marker(
                location=[float(row[7]), float(row[8])],
                popup=row[1],
                icon=None,
            ).add_to(marker_cluster)

    # Добавляем встроенный стиль для скрытия элемента с классом leaflet-control-attribution
    map_with_markers.get_root().html.add_child(folium.Element("<style>.leaflet-control-attribution { display: none; }</style>"))

    return map_with_markers


def webAppKeyboard(url):
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    webAppTest = types.WebAppInfo(url)
    one_butt = types.InlineKeyboardButton(text="для мобильных устройств", web_app=webAppTest)
    two_butt = types.InlineKeyboardButton(text="ПК", url=url)
    keyboard.add(one_butt, two_butt)
    return keyboard  # возвращаем клавиатуру


def webAppKeyboard_jt(url):
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    webAppTest = types.WebAppInfo(url)
    button = types.InlineKeyboardButton(text="открыть секретную страницу", web_app=webAppTest)

    keyboard.add(button)
    return keyboard  # возвращаем клавиатуру


async def filter_and_send_data(message, filter_func, command):
    goroda_data = load_goroda_data()
    headers = ['Наименование населенного пункта', 'Население 2010', 'Население 2020', 'Сотовая связь', 'Программа']
    filtered_data = [headers]
    filtered_goroda_data = []

    for row in goroda_data:
        if filter_func(row):
            filtered_row = [row[i] if i < len(row) else '' for i in [1, 2, 5, 3, 11]]
            filtered_data.append(filtered_row)
            filtered_goroda_data.append(row)

    # Создать карту с маркерами
    map_with_markers = await create_map_with_markers(filtered_goroda_data)
    map_filename = f"{command}_map.html"
    map_with_markers.save(map_filename)

    # Конвертировать данные в формат Excel и отправить
    buffer = convert_to_excel(filtered_data)
    filename = f"{command}.xlsx"
    with open(filename, "wb") as excel_file:
        excel_file.write(buffer.getvalue())

    with open(filename, "rb") as excel_file:
        document = InputFile(excel_file)
        await message.answer_document(document=document, caption="Список населенных пунктов")

    os.remove(filename)

    # Отправить файл с картой
    # with open(map_filename, "rb") as map_file:
    #    document = InputFile(map_file)
    #  bot.send_document(message.chat.id, document=document, caption=map_filename)

    os.remove(map_filename)
    url = f"https://rejoller.pythonanywhere.com/{command}"
    await message.answer("Чтобы посмотреть карту, нажмите кнопку ниже", reply_markup=webAppKeyboard(url))


def filter_2g(row):
    pattern = r"\b(2G|3G|4G)\b"
    result = re.findall(pattern, row[3])
    # filename = "2G.xlsx"
    return bool(result)


def filter_3g(row):
    pattern = r"\b(3G|4G)\b"
    result = re.findall(pattern, row[3])
    return bool(result)


def filter_4g(row):
    return "4G" in row[3]


def filter_nomobile(row):
    return row[3] == "-"


@dp.message_handler(commands=['2g'])
async def handle_2g_command(message: types.Message):
    log_user_data_from_message(message)
    await message.answer('Загружаю данные')
    await filter_and_send_data(message, filter_2g, "2G")


@dp.message_handler(commands=['3g'])
async def handle_3g_command(message: types.Message):
    log_user_data_from_message(message)
    await message.answer('Загружаю данные')
    await filter_and_send_data(message, filter_3g, "3G")


@dp.message_handler(commands=['4g'])
async def handle_4g_command(message: types.Message):
    await message.answer('Загружаю данные')
    await filter_and_send_data(message, filter_4g, "4G")


@dp.message_handler(commands=['nomobile'])
async def handle_nomobile_command(message: types.Message):
    await message.answer('Загружаю данные')
    await filter_and_send_data(message, filter_nomobile, "nomobile")


@dp.message_handler(commands=['jt'])
async def handle_jt_command(message: types.Message):
    log_user_data_from_message(message)

    url = f"https://fantastic-engine.vercel.app/"
    await message.answer("😁")
    await asyncio.sleep(3)
    await message.answer("🤭", reply_markup=webAppKeyboard_jt(url))


def load_fp_data():
    # Загружаем данные из Google Sheets
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID_1, range='ФАП!A1:M55').execute()
    rows = result.get('values', [])

    fp_data = []
    fp_headers = ['Адрес', 'Тип подключения', 'Скорость', 'Контакты', 'Дата подписания']

    for row in rows:
        # Выгружаем непустые строки и столбцы B, C, D, F, G, H, I
        if any(row) and "Исключение" not in row:  # Проверяем, что строка не пустая и не содержит "Исключ"
            filled_row = [row[i] if i < len(row) else '' for i in [1, 2, 3, 5, 6, 7, 8]]
            fp_data.append(filled_row)

    return fp_data, fp_headers

#5


async def load_aggoics_data():
    # Загружаем данные из Google Sheets
    gc = gspread.AsyncioAuthManager(credentials=creds)
    await gc.authorize()
    sheet = await gc.open_by_key(SPREADSHEET_ID_1).worksheet('АгГОиЧС')
    rows = await sheet.get_all_values()

    aggoics_data = []
    aggoics_headers = ['Муниципальное образование', 'Наименование населенного пункта', 'Адрес', 'Тип подключения', 'Наименование учреждения', 'Скорость']

    for row in rows:
        # Выгружаем непустые строки и столбцы D, E, F, G, I, J
        if any(row):  # Проверяем, что строка не пустая
            filled_row = [row[i] if i < len(row) else '' for i in [3, 4, 5, 6, 8, 9]]
            aggoics_data.append(filled_row)

    return aggoics_data, aggoics_headers


async def load_otpusk_data():
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    sheet = await spreadsheet.worksheet('otpusk')
    rows = await sheet.get('A1:F100')
    return rows




def get_employees_on_vacation(otpusk_data, days_ahead=3):
    today = datetime.today().date()
    future_vacation_start = today + timedelta(days=days_ahead)
    employees_on_vacation = []
    employees_starting_vacation_soon = []

    for row_idx, row in enumerate(otpusk_data):
        if row_idx == 0:  # пропустить заголовки таблицы
            continue
        if len(row) >= 5:
            try:
                start_date = datetime.strptime(row[3], "%d.%m.%Y").date()
                end_date = datetime.strptime(row[4], "%d.%m.%Y").date()

                if start_date <= today <= end_date:
                    employees_on_vacation.append(row)

                if today < start_date <= future_vacation_start:
                    employees_starting_vacation_soon.append(row)

            except ValueError:
                pass  # игнорировать строки с неправильным форматом даты

    return employees_on_vacation, employees_starting_vacation_soon

from PIL import Image, ImageDraw, ImageFont
import matplotlib.pyplot as plt


async def create_individual_radar_chart(chat_id, data_df, title):
    print("create_individual_radar_chart called with data:", data_df)

    # Создайте новое изображение с белым фоном
    img_width, img_height = 1000, 600
    img = Image.new('RGB', (img_width, img_height), 'white')
    draw = ImageDraw.Draw(img)

    # Добавьте заголовок
    title_font_path = "/home/rejoller/mysite/static/ofont.ru_Geologica.ttf"
    title_font = ImageFont.truetype(title_font_path, 30)
    text_font = ImageFont.truetype(title_font_path, 18)

    title_bbox = draw.textbbox((0,0), title, font=title_font)
    title_width, title_height = title_bbox[2] - title_bbox[0], title_bbox[3] - title_bbox[1]
    draw.text(((img_width - title_width) // 2, 20), title, fill="black", font=title_font)

    # Загрузите логотипы и уменьшите их
    logo_paths = [
        '/home/rejoller/mysite/static/tele2_1.png',
        '/home/rejoller/mysite/static/megafon_1.png',
        '/home/rejoller/mysite/static/beeline_1.png',

        '/home/rejoller/mysite/static/mts_1.png',
    ]

    logos = []
    resize_factors = [0.1, 0.1, 0.1*2, 0.1/3] # Уменьшаем МТС в 3 раза меньше и увеличиваем Билайн в 2 раза больше
    for i, path in enumerate(logo_paths):
        logo = Image.open(path)
        logo_width, logo_height = logo.size
        logos.append(logo.resize((int(logo_width * resize_factors[i]), int(logo_height * resize_factors[i]))))

    # Добавьте логотипы
    column_width = img_width // 4
    for i, logo in enumerate(logos):
        x = column_width * i + (column_width - logo.width) // 2
        y = 100
        if i in [1, 2]:  # индексы для Билайн и Мегафон
            # Создаем отдельное изображение для наложения
            logo_img = Image.new('RGBA', (img_width, img_height), (255, 255, 255, 0))
            logo_img.paste(logo, (x, y))

            # Накладываем логотип на основное изображение
            img = Image.alpha_composite(img.convert('RGBA'), logo_img).convert('RGB')
        else:
            # Для других логотипов просто вставляем их
            img.paste(logo, (x, y))

    # Создаем новый объект draw для текущего изображения
    draw = ImageDraw.Draw(img)

    # Сформулируйте текст для каждой строки в data_df и добавьте его на график
    operator_columns = [
        ('Уровень_Tele2', 'Качество_Tele2'),
        ('Уровень_Megafon', 'Качество_Megafon'),
        ('Уровень_Beeline', 'Качество_Beeline'),
        ('Уровень_MTS', 'Качество_MTS')
    ]

    y_start = y + logos[0].height + 20
    y_step = 20

    for idx, row_series in data_df.iterrows():
        for i, (level_column, quality_column) in enumerate(operator_columns):
            # Проверка на наличие данных для каждого оператора
            if pd.notnull(row_series[level_column]) or pd.notnull(row_series[quality_column]):
                text = f"{row_series.get(level_column, 'Нет данных')} {row_series.get(quality_column, 'Нет данных')}"
            else:
                text = "Нет данных"

            x = column_width * i + (column_width - logos[i].width) // 2  # Исправляем позиционирование текста
            y_text = y_start + idx * y_step
            print(f"Drawing text at ({x}, {y_text}): {text}")
            draw.text((x, y_text), text, fill="black", font=text_font)

    # Сохраните и отправьте изображение
    temp_file_path = "temp_survey_result.png"
    img.save(temp_file_path)

    # Отправляем изображение пользователю
    await bot.send_photo(chat_id, open(temp_file_path, 'rb'))

    # Удаляем временный файл
    os.remove(temp_file_path)
















def create_pie_chart(yes_count, no_count, filename):
    labels = ['Есть', 'Нет']
    sizes = [yes_count, no_count]
    colors = ['#2ecc71', '#e74c3c']

    # Создайте объект figure с заданными размерами (ширина, высота) в дюймах
    plt.figure(figsize=(2, 2))  # Здесь 2.5 дюйма - это ширина и высота диаграммы

    plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
    plt.axis('equal')
    plt.savefig(filename, dpi=400, bbox_inches='tight')  # Установите разрешение (DPI) изображения и параметр bbox_inches
    plt.clf()


def create_bar_chart(data, filename):
    labels, yes_values, no_values = zip(*data)

    total_values = [yes + no for yes, no in zip(yes_values, no_values)]
    yes_percentages = [yes / total * 100 if total != 0 else 0 for yes, total in zip(yes_values, total_values)]
    no_percentages = [no / total * 100 if total != 0 else 0 for no, total in zip(no_values, total_values)]

    labels = labels[1:]
    yes_percentages = yes_percentages[1:]
    no_percentages = no_percentages[1:]

    y = np.arange(len(labels))
    width = 0.6
    colors = ['#2ecc71', '#e74c3c']

    fig, ax = plt.subplots(figsize=(12, 16), dpi=300)  # Устанавливаем размер и DPI изображения
    rects1 = ax.barh(y, yes_percentages, label='Процент подключенных услуг', color=colors[0], align='center')
    rects2 = ax.barh(y, no_percentages, label='Процент не подключенных услуг', left=yes_percentages, color=colors[1],
                     align='center')

    ax.set_title('Подключение к ТОРКНД в Красноярском крае')
    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.legend()

    xmin = 0
    xmax = 100
    ax.set_xlim([xmin, xmax])

    def autolabel(rects, labels):
        for rect, label in zip(rects, labels):
            width = rect.get_width()
            ax.annotate('{:.1f}%'.format(label),
                        xy=(width, rect.get_y() + rect.get_height() / 2),
                        xytext=(3, 0),
                        textcoords="offset points",
                        ha='left', va='center')

    autolabel(rects1, yes_percentages)
    # autolabel(rects2, no_percentages)

    plt.tight_layout()
    plt.savefig(filename)
    plt.close()

'''
async def search_szofed_values(column_4_value):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('szofed!A1:M2412')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if column_4_value.lower() == row[0].lower()]
    return found_values
'''
async def search_szofed_values(column_4_value, spreadsheet):
    result = await spreadsheet.values_batch_get('szofed!A1:M2412')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if column_4_value.lower() == row[0].lower()]
    return found_values



@dp.message_handler(commands=['aggoics'])
async def handle_aggoics_command(message: types.Message):
    await message.answer('Загружаю данные')
    aggoics_data, aggoics_headers = await load_aggoics_data()
    await filter_and_send_data(message, aggoics_data, aggoics_headers)


@dp.message_handler(commands=['otpusk'])
async def handle_otpusk_command(message: types.Message):
    await message.answer('Загружаю данные')
    otpusk_data = await load_otpusk_data()
    await filter_and_send_data(message, otpusk_data)


async def filter_and_send_data(message: types.Message, data, headers=None):
    if headers:
        filtered_data = [headers] + data
    else:
        filtered_data = data

    filtered_data = [list(map(str, row)) for row in filtered_data]

    # Отправляем данные частями, чтобы не превысить лимит сообщения
    for i in range(0, len(filtered_data), 10):
        chunk = filtered_data[i:i+10]
        text = '\n'.join(['\t'.join(row) for row in chunk])
        await message.answer(text, parse_mode=types.ParseMode.MARKDOWN)


@dp.message_handler(commands=['employees_vacation'])
async def handle_employees_vacation_command(message: types.Message):
    await message.answer('Загружаю данные')
    otpusk_data = await load_otpusk_data()
    employees_on_vacation, employees_starting_vacation_soon = get_employees_on_vacation(otpusk_data)

    if employees_on_vacation:
        await message.answer('Сотрудники, находящиеся в отпуске:')
        await message.answer('\n'.join(['\t'.join(row) for row in employees_on_vacation]), parse_mode=types.ParseMode.MARKDOWN)
    else:
        await message.answer('Сотрудников в отпуске нет.')

    if employees_starting_vacation_soon:
        await message.answer('Сотрудники, начинающие отпуск в ближайшие дни:')
        await message.answer('\n'.join(['\t'.join(row) for row in employees_starting_vacation_soon]), parse_mode=types.ParseMode.MARKDOWN)
    else:
        await message.answer('Сотрудников, начинающих отпуск в ближайшие дни, нет.')


@dp.message_handler(commands=['pie_chart'])
async def handle_pie_chart_command(message: types.Message):
    await message.answer('Создаю круговую диаграмму')
    create_pie_chart(20, 80, 'pie_chart.png')
    with open('pie_chart.png', 'rb') as photo:
        await message.answer_photo(photo)


@dp.message_handler(commands=['bar_chart'])
async def handle_bar_chart_command(message: types.Message):
    await message.answer('Создаю гистограмму')
    data = [
        ('Район 1', 5, 10),
        ('Район 2', 10, 15),
        ('Район 3', 20, 5),
        ('Район 4', 30, 25),
        ('Район 5', 50, 20),
    ]
    create_bar_chart(data, 'bar_chart.png')
    with open('bar_chart.png', 'rb') as photo:
        await message.answer_photo(photo)

async def get_authorized_client_and_spreadsheet():
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    return gc, spreadsheet


@dp.message_handler(commands=['search_szofed'])
async def handle_search_szofed_command(message: types.Message):
    await message.answer('Поиск данных')
    gc, spreadsheet = await get_authorized_client_and_spreadsheet()
    found_values = await search_szofed_values('some_value', spreadsheet)
    if found_values:
        await message.answer('Найденные значения:')
        await message.answer('\n'.join(['\t'.join(row) for row in found_values]), parse_mode=types.ParseMode.MARKDOWN)
    else:
        await message.answer('Значения не найдены.')


'''
@dp.message_handler(commands=['search_szofed'])
async def handle_search_szofed_command(message: types.Message):
    await message.answer('Поиск данных')
    found_values = await search_szofed_values('some_value')
    if found_values:
        await message.answer('Найденные значения:')
        await message.answer('\n'.join(['\t'.join(row) for row in found_values]), parse_mode=types.ParseMode.MARKDOWN)
    else:
        await message.answer('Значения не найдены.')

async def search_espd_values(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('espd!A1:AL1466')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if query.lower() == row[0].lower()]
    return found_values

async def search_szoreg_values(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('szoreg!A1:Q1700')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if query.lower() == row[0].lower()]
    return found_values
'''
#nтест



async def search_espd_values(query, spreadsheet):
    try:
        # Сгенерировать ключ для кэширования на основе запроса
      #  cache_key = f"espd_values:{query.lower()}"


        # Попытаться получить данные из кэша
     #   cached_data = await storage.get_data(chat=cache_key)
     #   if cached_data:
       #     cached_data = json.loads(cached_data)  # Десериализация данных

    #        return cached_data



        # Если данных в кэше нет, продолжаем поиск в таблице
        result = await spreadsheet.values_batch_get('espd!A1:AL1466')
        rows = result.get('valueRanges', [])[0].get('values', [])
        found_values = [row for row in rows if query.lower() == row[0].lower()]

        # Сохранить результаты в кэше перед возвратом
      #  await storage.set_data(chat=cache_key, data=json.dumps(found_values))



        return found_values
    except Exception as e:
        print("An error occurred during search_espd_values:", e)
        return None




cached_szoreg_values = None

async def load_szoreg_values(spreadsheet):
    global cached_szoreg_values
    try:
        if cached_szoreg_values is None:
            result = await spreadsheet.values_batch_get('szoreg!A1:Q10000')
            rows = result.get('valueRanges', [])[0].get('values', [])
            cached_szoreg_values = rows
        return cached_szoreg_values
    except Exception as e:
        print("An error occurred during loading szoreg_values:", e)
        return None

async def search_szoreg_values(query, spreadsheet):
    try:
        # Загрузить данные, если они еще не загружены
        szoreg_values = await load_szoreg_values(spreadsheet)

        # Найти нужные значения в загруженных данных
        found_values = [row for row in szoreg_values if query.lower() == row[0].lower() and (len(row) < 12 or row[11] != 'Исключение')]

        return found_values
    except Exception as e:
        print("An error occurred during search_szoreg_values:", e)
        return None



'''

async def search_szoreg_values(query, spreadsheet):
    try:
        # Сгенерировать ключ для кэширования на основе запроса
     #   cache_key = f"szoreg_values:{query.lower()}"


        # Попытаться получить данные из кэша
     #   cached_data = await storage.get_data(chat=cache_key)
   #     if cached_data:
       #     cached_data = json.loads(cached_data)  # Десериализация данных

       #     return cached_data




        result = await spreadsheet.values_batch_get('szoreg!A1:Q10000')
        rows = result.get('valueRanges', [])[0].get('values', [])
        #found_values = [row for row in rows if query.lower() == row[0].lower()]
        found_values = [row for row in rows if query.lower() == row[0].lower() and (len(row) < 12 or row[11] != 'Исключение')]
        print('found_values:', found_values)
        # Сохранить результаты в кэше перед возвратом
    #    await storage.set_data(chat=cache_key, data=json.dumps(found_values))



        return found_values
    except Exception as e:
        print("An error occurred during search_szoreg_values:", e)
        return None
'''



async def search_mszu_mo(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('МСЗУ-ОМСУ (тест)!A1:P3200')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if query.lower() == row[0].lower()]
    return found_values
'''
async def search_yandex_2023_values(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('2023!A3:P50')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if query.lower() == row[0].lower()]
    return found_values

async def search_in_pokazatel_504p(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('показатель 504-п!A1:K1719')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if query.lower() == row[0].lower()]
    return found_values

async def search_in_ucn2(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('УЦН 2.0 (2023)!A1:K800')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if query.lower() == row[0].lower()]
    return found_values

async def search_schools_values(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('Школы!A1:U1500')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if query.lower() == row[0].lower()]
    return found_values

async def get_votes_data():
    try:
        agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
        gc = await agcm.authorize()
        spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
        result = await spreadsheet.values_batch_get('votes!A2:C2000')
        rows = result.get('valueRanges', [])[0].get('values', [])
        return rows
    except Exception as e:
        print("An error occurred while getting votes data:", e)
        raise e
'''


async def search_yandex_2023_values(query, spreadsheet):
    try:
        # Сгенерировать ключ для кэширования на основе запроса
        #cache_key = f"yandex_2023_values:{query.lower()}"


        # Попытаться получить данные из кэша
       # cached_data = await storage.get_data(chat=cache_key)
       # if cached_data:
       #     cached_data = json.loads(cached_data)  # Десериализация данных

      #     return cached_data


        # Если данных в кэше нет, продолжаем поиск в таблице
        result = await spreadsheet.values_batch_get('2023!A3:P50')
        rows = result.get('valueRanges', [])[0].get('values', [])
        found_values = [row for row in rows if query.lower() == row[0].lower()]

        # Сохранить результаты в кэше перед возвратом
      #  await storage.set_data(chat=cache_key, data=json.dumps(found_values))



        return found_values
    except Exception as e:
        print("An error occurred during search_yandex_2023_values:", e)
        return None





# Глобальная переменная для хранения всего диапазона данных
cached_pokazatel_504p_values = None

async def load_pokazatel_504p_values(spreadsheet):
    global cached_pokazatel_504p_values
    try:
        if cached_pokazatel_504p_values is None:
            result = await spreadsheet.values_batch_get('показатель 504-п!A1:K1719')
            rows = result.get('valueRanges', [])[0].get('values', [])
            cached_pokazatel_504p_values = rows
        return cached_pokazatel_504p_values
    except Exception as e:
        print("An error occurred during loading pokazatel_504p_values:", e)
        return None

async def search_in_pokazatel_504p(query, spreadsheet):
    try:
        # Загрузить данные, если они еще не загружены
        pokazatel_504p_values = await load_pokazatel_504p_values(spreadsheet)

        # Найти нужные значения в загруженных данных
        found_values = [row for row in pokazatel_504p_values if query.lower() == row[0].lower()]

        return found_values
    except Exception as e:
        print("An error occurred during search_in_pokazatel_504p:", e)
        return []


'''

async def search_in_pokazatel_504p(query, spreadsheet):
    try:
        # Сгенерировать ключ для кэширования на основе запроса
       # cache_key = f"pokazatel_504p_values:{query.lower()}"


        # Попытаться получить данные из кэша
      #  cached_data = await storage.get_data(chat=cache_key)
       # if cached_data:
      #      cached_data = json.loads(cached_data)  # Десериализация данных

       #     return cached_data



        # Если данных в кэше нет, продолжаем поиск в таблице
        result = await spreadsheet.values_batch_get('показатель 504-п!A1:K1719')
        rows = result.get('valueRanges', [])[0].get('values', [])
        found_values = [row for row in rows if query.lower() == row[0].lower()]

        # Сохранить результаты в кэше перед возвратом
     #   await storage.set_data(chat=cache_key, data=json.dumps(found_values))



        return found_values
    except Exception as e:
        print("An error occurred during search_in_pokazatel_504p:", e)
        return []
'''


# Глобальная переменная для хранения всего диапазона данных
cached_ucn2_values = None

async def load_ucn2_values(spreadsheet):
    global cached_ucn2_values
    try:
        if cached_ucn2_values is None:
            result = await spreadsheet.values_batch_get('УЦН 2.0 (2023)!A1:K800')
            rows = result.get('valueRanges', [])[0].get('values', [])
            cached_ucn2_values = rows
        return cached_ucn2_values
    except Exception as e:
        print("An error occurred during loading ucn2_values:", e)
        return None

async def search_in_ucn2(query, spreadsheet):
    try:
        # Загрузить данные, если они еще не загружены
        ucn2_values = await load_ucn2_values(spreadsheet)

        # Найти нужные значения в загруженных данных
        found_values = [row for row in ucn2_values if query.lower() == row[0].lower()]

        return found_values
    except Exception as e:
        print("An error occurred during search_in_ucn2:", e)
        return None


'''
async def search_in_ucn2(query, spreadsheet):
    try:
        # Сгенерировать ключ для кэширования на основе запроса
      #  cache_key = f"ucn2_values:{query.lower()}"


        # Попытаться получить данные из кэша
       # cached_data = await storage.get_data(chat=cache_key)
       # if cached_data:
          #  cached_data = json.loads(cached_data)  # Десериализация данных

           # return cached_data



        # Если данных в кэше нет, продолжаем поиск в таблице
        result = await spreadsheet.values_batch_get('УЦН 2.0 (2023)!A1:K800')
        rows = result.get('valueRanges', [])[0].get('values', [])
        found_values = [row for row in rows if query.lower() == row[0].lower()]

        # Сохранить результаты в кэше перед возвратом
       # await storage.set_data(chat=cache_key, data=json.dumps(found_values))



        return found_values
    except Exception as e:
        print("An error occurred during search_in_ucn2:", e)
        return None
'''

async def search_schools_values(query, spreadsheet):
    try:
        # Сгенерировать ключ для кэширования на основе запроса
      #  cache_key = f"schools_values:{query.lower()}"


        # Попытаться получить данные из кэша
        #cached_data = await storage.get_data(chat=cache_key)
       # if cached_data:
        #   cached_data = json.loads(cached_data)  # Десериализация данных

         #   return cached_data



        # Если данных в кэше нет, продолжаем поиск в таблице
        result = await spreadsheet.values_batch_get('Школы!A1:U1500')
        rows = result.get('valueRanges', [])[0].get('values', [])
        found_values = [row for row in rows if query.lower() == row[0].lower()]

        # Сохранить результаты в кэше перед возвратом
       # await storage.set_data(chat=cache_key, data=json.dumps(found_values))



        return found_values
    except Exception as e:
        print("An error occurred during search_schools_values:", e)
        return None


async def get_votes_data(spreadsheet):
    try:
        result = await spreadsheet.values_batch_get('votes!A2:D2000')
        rows = result.get('valueRanges', [])[0].get('values', [])
        return rows
    except Exception as e:
        print("An error occurred while getting votes data:", e)
        raise e



async def search_in_results(query, spreadsheet):
    try:
        # Сгенерировать ключ для кэширования на основе запроса
        #cache_key = f"results:{query.lower()}"


        # Попытаться получить данные из кэша
        #cached_data = await storage.get_data(chat=cache_key)
        #if cached_data:
         #   cached_data = json.loads(cached_data)  # Десериализация данных

         #   return cached_data


        # Если данных в кэше нет, продолжаем поиск в таблице
        result = await spreadsheet.values_batch_get('Результаты опроса!A1:N')
        rows = result.get('valueRanges', [])[0].get('values', [])
        found_values = [row for row in rows if query.lower() == row[5].lower()]

        # Сохранить результаты в кэше перед возвратом
       # await storage.set_data(chat=cache_key, data=json.dumps(found_values))


        return found_values
    except Exception as e:
        print(f"An error occurred during search_in_results: {e}")
        return None


'''
async def get_votes_tanya_data():
    try:
        agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
        gc = await agcm.authorize()
        spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
        result = await spreadsheet.values_batch_get('Голосование Таня!A2:C2000')
        rows = result.get('valueRanges', [])[0].get('values', [])
        return rows
    except Exception as e:
        print("An error occurred while getting votes data:", e)
        raise e

'''









import traceback

import tempfile

import shutil

@dp.message_handler(commands=['votes'])
async def send_votes(message: types.Message):
    try:
        gc, spreadsheet = await get_authorized_client_and_spreadsheet()
        data = await get_votes_data(spreadsheet)
        excel_data = create_excel_file_2(data)  # убрали headers здесь
        await log_user_data_from_message(message)
        # Сохраняем данные Excel во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp:
            temp.write(excel_data.read())
            temp_filename = temp.name

        # Переименовываем файл перед отправкой
        final_filename = "Голосование УЦН 2_0 2024.xlsx"
        shutil.move(temp_filename, final_filename)

        # Отправляем файл
        with open(final_filename, "rb") as temp:
            await bot.send_document(message.chat.id, temp, caption='Информация о голосовании УЦН 2.0 2024')

        # Удаляем файл после отправки
        os.remove(final_filename)

    except Exception as e:
        tb = traceback.format_exc()  # Получить трассировку стека
        print("An error occurred while handling /votes:", tb)  # Печатает трассировку стека
        await message.reply(f'Произошла ошибка при обработке вашего запроса: {e}\n{tb}')  # Включает ошибку и трассировку стека в ответ пользователю




@dp.message_handler(commands=['Tanya_dushnila_ucn'])
async def send_votes(message: types.Message):
    try:
        data = await get_votes_tanya_data()
        excel_data = create_excel_file_2(data)  # убрали headers здесь
        await log_user_data_from_message(message)
        # Сохраняем данные Excel во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp:
            temp.write(excel_data.read())
            temp_filename = temp.name

        # Переименовываем файл перед отправкой
        final_filename = "Таня УЦН 2_0 2024.xlsx"
        shutil.move(temp_filename, final_filename)

        # Отправляем файл
        with open(final_filename, "rb") as temp:
            await bot.send_document(message.chat.id, temp, caption='Информация о голосовании УЦН 2.0 2024 для Тани🐭')

        # Удаляем файл после отправки
        os.remove(final_filename)

    except Exception as e:
        tb = traceback.format_exc()  # Получить трассировку стека
        print("An error occurred while handling /votes:", tb)  # Печатает трассировку стека
        await message.reply(f'Произошла ошибка при обработке вашего запроса: {e}\n{tb}')  # Включает ошибку и трассировку стека в ответ пользователю






@dp.message_handler(commands=['start'])
async def handle_start(message: types.Message):
    user_first_name = message.from_user.first_name
    await message.answer(
        f'Привет, {user_first_name}!\nЯ бот который может поделиться с тобой информацией о связи в Красноярском крае. Для этого введи название населенного пункта или муниципального образования (например "Курагино" или "Абанский")\nЧтобы узнать информацию о сотовой связи, выбери /2g /3g или /4g. Чтобы получить информацию о населенных пунктах без сотовой связи жми /nomobile \n\n'
        'Чтобы узнать о подключении к ТОРКНД, введи сообщение "тор" и наименование муниципального образования. '
        'Например, "тор Енисейский".\n'
        'Если нужна статистика по всему краю, жми /knd_kraj\n\n'
        'Чтобы узнать, кто сегодня в отпуске, жми /otpusk\n'

        'Если остались вопросы, пиши @rejoller.')


def preprocess_rows(rows):
    preprocessed_rows = []
    for row in rows:
        if len(row) > 1:
            lemmatized_keywords = {token.lemma_ for token in nlp(row[1].lower()) if not token.is_stop and not token.is_punct}
            preprocessed_rows.append((row, lemmatized_keywords))
    return preprocessed_rows


async def check_mszu_column_b(user_message):
    async with aiohttp.ClientSession() as session:
        async with session.get(f"https://sheets.googleapis.com/v4/spreadsheets/{SPREADSHEET_ID_1}/values/МСЗУ!A1:P150", headers={"Authorization": f"Bearer {creds.token}"}) as response:
            result = await response.json()
            values = result.get('values', [])
            preprocessed_values = preprocess_rows(values)
            user_doc = nlp(user_message.lower())
            user_keywords = {token.lemma_ for token in user_doc if not token.is_stop and not token.is_punct}
            matching_rows = [row for row, lemmatized_keywords in preprocessed_values if user_keywords == lemmatized_keywords]

            if not matching_rows:
                matching_rows = [row for row, lemmatized_keywords in preprocessed_values if user_keywords.intersection(lemmatized_keywords)]

            return matching_rows if len(matching_rows) > 0 else None
#7
def ngrams(sequence, n):
    return list(nltk.ngrams(sequence, n))

def jaccard_similarity(a, b):
    a_set = set(a)
    b_set = set(b)
    return len(a_set.intersection(b_set)) / len(a_set.union(b_set))

def preprocess_rows_2(rows):
    preprocessed_rows = []
    for row in rows:
        if len(row) > 1:  # Проверка наличия данных в столбце B
            b_column_value = row[1]  # Индекс 1 соответствует столбцу B
            row_doc = nlp(b_column_value.lower())
            row_keywords = {}
            for i, token in enumerate(row_doc):
                if not token.is_stop and not token.is_punct and len(token) > 2:
                    weight = 2 if i + 1 < len(row_doc) and row_doc[i + 1].text.lower() in ["сельсовет", "район", "округ", "муниципальный округ"] else 1
                    row_keywords[token.lemma_] = weight
            preprocessed_rows.append((row, row_keywords))
    return preprocessed_rows

def weighted_keyword_match(user_keywords, row_keywords):
    user_keyword_set = set(user_keywords.keys())
    row_keyword_set = set(row_keywords.keys())

    intersection = user_keyword_set.intersection(row_keyword_set)
    union = user_keyword_set.union(row_keyword_set)

    if not union:
        return 0

    weighted_intersection_sum = sum([user_keywords.get(k, 0) * row_keywords.get(k, 0) for k in intersection])
    weighted_union_sum = sum([user_keywords.get(k, 0) for k in union]) + sum([row_keywords.get(k, 0) for k in union]) - weighted_intersection_sum

    return weighted_intersection_sum / weighted_union_sum

def check_mszu_mo(user_message):
    user_doc = nlp(user_message.lower())
    user_keywords = {
        token.lemma_: 2 if token.text.lower() in ["сельсовет", "район", "округ", "муниципальный округ"] else 1
        for token in user_doc
        if not token.is_stop and not token.is_punct and len(token) > 2
    }

    # Создание сервиса для доступа к API Google Sheets
    service = build('sheets', 'v4', credentials=creds)

    index_range_name = 'mszuindex!A1:C500'
    main_range_name = 'МСЗУ-ОМСУ (тест)!A1:T3200'

    index_result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID_1, range=index_range_name).execute()
    main_result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID_1, range=main_range_name).execute()

    index_values = index_result.get('values', [])
    main_values = main_result.get('values', [])

    preprocessed_index_values = preprocess_rows_2(index_values)

    matching_rows = []
    for index_row, index_keywords in preprocessed_index_values:
        if weighted_keyword_match(user_keywords, index_keywords) >= 0.75:
            index_c_value = index_row[2]  # Получаем значение из столбца C таблицы mszuindex
            # Находим соответствующие строки в таблице "МСЗУ-ОМСУ (тест)" путем сравнения столбца J с index_c_value
            matched_rows = [row for row in main_values if row[9] == index_c_value]
            matching_rows.extend(matched_rows)

    return matching_rows

def preprocess_rows_3(rows):
    preprocessed_rows = []
    for row in rows:
        if len(row) > 5:  # Проверка наличия данных в столбце F
            f_column_value = row[5]  # Индекс 5 соответствует столбцу F
            row_doc = nlp(f_column_value.lower())
            row_keywords = {}
            for i, token in enumerate(row_doc):
                if not token.is_stop and not token.is_punct and len(token) > 2:
                    weight = 2 if i + 1 < len(row_doc) and row_doc[i + 1].text.lower() in ["район", "го", "мо", "округ", "муниципальный округ"] else 1
                    row_keywords[token.lemma_] = weight
            preprocessed_rows.append((row, row_keywords))
    return preprocessed_rows


#8
async def check_mszu_mo_2(user_message):
    user_doc = nlp(user_message.lower())
    user_keywords = {
        token.lemma_: 2 if token.text.lower() in ["сельсовет", "район", "округ", "муниципальный округ"] else 1
        for token in user_doc
        if not token.is_stop and not token.is_punct and len(token) > 2
    }

    index_range_name = 'mszuindex!A1:G500'
    main_range_name = 'МСЗУ-ОМСУ (тест)!A1:T3200'

    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)

    index_result = await spreadsheet.values_get(index_range_name)
    main_result = await spreadsheet.values_get(main_range_name)

    index_values = index_result.get('values', [])
    main_values = main_result.get('values', [])

    preprocessed_index_values = preprocess_rows_3(index_values)

    matching_rows = []
    for index_row, index_keywords in preprocessed_index_values:
        if weighted_keyword_match(user_keywords, index_keywords) >= 0.75:
            index_g_value = index_row[6]  # Получаем значение из столбца G таблицы mszuindex
            # Находим соответствующие строки в таблице "МСЗУ-ОМСУ (тест)" путем сравнения столбца K с index_g_value
            matched_rows = [row for row in main_values if row[10] == index_g_value]
            matching_rows.extend(matched_rows)

    return matching_rows
'''
@dp.message_handler()
async def process_message(message: types.Message):
    print ('async def process_message(message: types.Message):')
    user_message = message.text
    matching_rows = await check_mszu_mo_2(user_message)
    if matching_rows:
        for row in matching_rows:
            await message.reply(row)
'''


def send_request_to_openai_post(chat_id):
    #log_user_data_from_message(chat_id)
    global info_text_storage
    trimmed_info = info_text_storage[chat_id][:7000]
    messages = [
        {"role": "system", "content": "ты сотрудник пресс-службы министерства цифрового развития Красноярского края. Ты специализируешься на создании постов в социальных сетях о визитах министра Распопина Николая Александровича"},

        {"role": "assistant", "content": trimmed_info},

        {"role": "user", "content": "Создай креативный пост для социальной сети 'ВК', используя всю информацию по этому населенному пункту, добавляй эмодзи там где это необходимо и теги в конце поста. Скажи что планируются встерчи с жителями по вопросам качества предоставляемых услуг связи, а также  по вопросам реализации проектов цифровой трансформации с главой населенного пункта и главой района, если этот населенный пункт входит в состав какого-то района или муниципального округа. Если есть объекты СЗО, то также нужно указать что планируется визит в них, чтобы проверить как созданная инфраструктура используется и какую пользу приносит. Не обязательно чтобы всё было слово в слово, можешь перефразировать так как считаешь нужным и добавляй различные детали, также используй информацию о погоде. Визит планируется на следующей неделе. "}

    ]
    response = openai.ChatCompletion.create(
        model="gpt-4-turbo-preview",
        messages=messages,
        max_tokens=1500,
        n=1,
        temperature=0.6,
    )
    return response['choices'][0]['message']['content']

#9
async def send_request_to_openai(chat_id):
    #log_user_data_from_message(chat_id)
    global info_text_storage
    trimmed_info = info_text_storage[chat_id][:7000]
    messages = [{"role": "system", "content": "ты сотрудник министерства цифрового развития Красноярского края. Ты получаешь краткое содержание из обращений граждан. На основании полученной информации ты должен подготовить развернутый ответ. Избегай тавтологий и излишней формальности. стиль - официально деловой. Обращай внимание на наименование населенного пункта и на численность населения при формировании своего ответа, а также о программах реализуемых на территориях насленных пунктов (при наличии). Ориентируйся на шаблоны ниже, но при этом необходимо плавно склеивать части ответа между собой, чтобы в результате получился целостный ответ, который приятно читать. Вежливость 9 из 10. Если видишь что сотовая связь имеет уровень 3G или 4G и при этом хорошее качество, то не рассказывай про таксофоны - это неуместно. Ответ должен быть полностью сформирован для копипаста. запомни, не упоминай про наличие таксофонов если видишь что в населенном пункте есть 3G 🟢Хорошее или 4G 🟢Хорошее"

        "Численность населения берется после слов [👥Население (2020 г): ] и используется при реализации программы [УЦН (устранение цифрового неравенства)2.0]. Также следует отметить, что если населенный пункт участвует в УЦН 2.0, то говорить о других госпрограммах не следует. И наоборот."
        "Вводные слова с которых следует начинать предложения: [Также сообщаем, дополнительно сообщаем, следует отметить, учитывая изложенное, в соответствии с изложенным ]"
        "Информация о справочнике и о списке сотрудников в отпуске не должна попадать в ответ - это корпоративная информация"        },
        {"role": "user", "content":      "сообщение в котором  численность населения менее 100; таксофон = [1]:"

       " Абанский р-н, д. Борки (19.86°C 🌧️)"

"[👥Население (2010 г):] 83 чел."
"[👥Население (2020 г):] 71 чел."

"Сотовая связь:🔴отсутствует"
"Интернет: "

"Количество таксофонов: 1"},
        {"role": "assistant", "content": "В ответ на Ваше обращение по вопросу предоставления услуг связи на территории [населенный пункт] министерство цифрового развития Красноярского края (далее – министерство) сообщает следующее."
"По информации операторов связи, на территории [наименование населенного пункта] отсутствуют услуги подвижной радиотелефонной связи (далее – ПРТС) из-за значительной удаленности населенного пункта от инфраструктурных объектов связи, а также очень сложным рельефом местности. Инвестиционными планами операторов не предусмотрена установка объекта связи на территории [населенный пункт]"
"Министерством цифрового развития, связи и массовых коммуникаций Российской Федерации совместно с оператором связи ПАО «Ростелеком» реализуется федеральный проект «Устранение цифрового неравенства», в рамках которого в населенных пунктах с численностью населения 100-500 человек устанавливаются средства связи для оказания услуг ПРТС. По данным Всероссийской переписи населения 2020 году в [населенный пункт] проживает [👥Население (2020 г):] , что не позволяет принять участие в указанном проекте. "
"В рамках мероприятия «Субсидии бюджетам муниципальных образований на создание условий для обеспечения услугами связи малочисленных и труднодоступных населенных пунктов Красноярского края» государственной программы «Развитие информационного общества» субсидии бюджетам муниципальных образований предоставляются на основе конкурсного отбора."
"Согласно критериям отбора, обозначенных в п. 9 порядка предоставления и распределения субсидий бюджетам муниципальных образований Красноярского края на создание условий для обеспечения услугами связи малочисленных и труднодоступных населенных пунктов Красноярского края, утвержденным постановлением Правительства Красноярского края от 31.12.2019 № 791-п, проводится оценка каждой поданной от муниципальных образований заявки, при этом одним из наиболее весомых при подведении итогов является критерий «Стоимость организации услуг связи в расчете на одного жителя населенного пункта за год», рассчитываемый как отношение суммы коммерческого предложения к численности населения."
"Ввиду существенных финансовых затрат, связанных с организацией услуг связи, включая строительство ВОЛС и соответствующих объектов на территории [населенный пункт], резкий контраст с малой численностью жителей повлечет за собой уменьшение количества итоговых баллов и, как следствие, исключение населенного пункта из итогового состава субсидии."
"Тем не менее на данный момент имеется возможность совершать бесплатные звонки внутри страны на любые номера без использования телефонной карты с помощью таксофонов, расположенных на территории населенного пункта. Следует отметить, "
"что за международные соединения взимается соответствующая плата. Телефонные карты можно приобрести в офисах ООО «Телекомсервис»."},


         {"role": "user", "content":      "сообщение в котором  численность населения менее 100; таксофон = [0]:"

       " Абанский р-н, д. Борки (19.86°C 🌧️)"

"[👥Население (2010 г): ]83 чел."
"[👥Население (2020 г): ]71 чел."

"Сотовая связь:🔴отсутствует"
"Интернет: "

"Количество таксофонов: 0"},


{"role": "assistant", "content": "В ответ на Ваше обращение по вопросу предоставления услуг связи на территории [населенный пункт] министерство цифрового развития Красноярского края (далее – министерство) сообщает следующее."
"По информации операторов связи, на территории [наименование населенного пункта] отсутствуют услуги подвижной радиотелефонной связи (далее – ПРТС) из-за значительной удаленности населенного пункта от инфраструктурных объектов связи, а также очень сложным рельефом местности. Инвестиционными планами операторов не предусмотрена установка объекта связи на территории [населенный пункт]"
"Министерством цифрового развития, связи и массовых коммуникаций Российской Федерации совместно с оператором связи ПАО «Ростелеком» реализуется федеральный проект «Устранение цифрового неравенства», в рамках которого в населенных пунктах с численностью населения 100-500 человек устанавливаются средства связи для оказания услуг ПРТС. По данным Всероссийской переписи населения 2020 году в [населенный пункт] проживает [👥Население (2020 г):] , что не позволяет принять участие в указанном проекте. "
"В рамках мероприятия «Субсидии бюджетам муниципальных образований на создание условий для обеспечения услугами связи малочисленных и труднодоступных населенных пунктов Красноярского края» государственной программы «Развитие информационного общества» субсидии бюджетам муниципальных образований предоставляются на основе конкурсного отбора."
"Согласно критериям отбора, обозначенных в п. 9 порядка предоставления и распределения субсидий бюджетам муниципальных образований Красноярского края на создание условий для обеспечения услугами связи малочисленных и труднодоступных населенных пунктов Красноярского края, утвержденным постановлением Правительства Красноярского края от 31.12.2019 № 791-п, проводится оценка каждой поданной от муниципальных образований заявки, при этом одним из наиболее весомых при подведении итогов является критерий «Стоимость организации услуг связи в расчете на одного жителя населенного пункта за год» (пункт 2), рассчитываемый как отношение суммы коммерческого предложения к численности населения."
"Ввиду существенных финансовых затрат, связанных с организацией услуг связи, включая строительство ВОЛС и соответствующих объектов на территории [населенный пункт], резкий контраст с малой численностью жителей повлечет за собой уменьшение количества итоговых баллов и, как следствие, исключение населенного пункта из итогового состава субсидии."
"Обращаем Ваше внимание, что при необходимости министерством будет рассмотрен вопрос установки таксофона универсальных услуг связи на территории [населенный пункт], который позволяет совершать бесплатные звонки внутри страны на любые номера, за международные соединения взимается плата. "},

{"role": "user", "content":      "сообщение в котором  численность населения от 100 до 500:"
"[👥Население (2010 г.):] 220 чел."
"[👥Население (2020 г.):] 172 чел."

"Сотовая связь: 🔴отсутствует"
"Интернет: "

"Количество таксофонов: 0"

"Информация из таблицы УЦН 2.0 (2023):"

"[Голосов через Госусулуги: 145"
"Голосов почтой: 73"
"Итого голосов: 218"
"Проверка услуг: Нет голоса, нет ПД"
"Примечание: УЦН 2.0 на 2023"
"Попадение в квоту: да]"},






 {"role": "assistant", "content":       "В ответ на Ваше обращение по вопросу предоставления услуг связи в [населенный пункт] министерство цифрового развития Красноярского края сообщает следующее."
"На территории [населенный пункт] услуги подвижной радиотелефонной связи отсутствуют или присутствуют низкого качества из-за значительной удаленности населенного пункта от инфраструктурных объектов связи. Инвестиционные планы развития сетей связи не предусматривают установку оборудования на территории обозначенного населенного пункта."
"Следует отметить, что Министерством цифрового развития, связи и массовых коммуникаций Российской Федерации совместно с оператором ПАО «Ростелеком» реализуется федеральный проект «Устранение цифрового неравенства», в рамках которого в населенных пунктах с численностью населения 100-500 человек устанавливаются средства связи для оказания услуг подвижной радиотелефонной связи."
"[Далее пиши по следующему алгоритму,"
"[если присутствует [Информация из таблицы 2023], то пиши [В рамках государственной программы «Развитие информационного общества» согласно постановлению Правительства Красноярского края от 30.11.2022 № 1030-п «Об утверждении распределения в 2023 году субсидий бюджетам муниципальных образований Красноярского края на создание условий для обеспечения услугами связи малочисленных и труднодоступных населенных пунктов Красноярского края» администрации [наименование муниципального образования] в 2023 году выделена субсидия на организацию услуг ПРТС на территории [населенный пункт]. На данный момент уже заключено соглашение между министерством и администрацией района о предоставлении субсидии бюджету [наименование муниципального образование] на создание условий для обеспечения услугами связи малочисленных и труднодоступных населенных пунктов Красноярского края. Предварительный срок запуск услуг ПРТС до конца 2023 года."
"[если присутствует [Информация из таблицы УЦН 2.0 (2023):] и там [Попадение в квоту: да], то пиши [Министерством цифрового развития, связи и массовых коммуникаций Российской Федерации совместно с оператором связи ПАО «Ростелеком» реализуется федеральный проект «Устранение цифрового неравенства», в рамках которого в населенных пунктах с численностью населения 100-500 человек устанавливаются средства связи для оказания услуг ПРТС. Так по итогам онлайн-голосования в ноябре 2022 года [наименование населенного пункта] включена в перечень населённых пунктов, планируемых к подключению услуг в срок до []."

"Если в [Информация из таблицы УЦН 2.0 (2023):] и [Попадение в квоту: ] =  , то пиши  [Перечень населенных пунктов формируется согласно итогам онлайн-голосования на портале государственных услуг, так по итогам голосования в ноябре 2022 года [населенный пункт] набрал [Итого голосов:] голосов, занимая ___ место при квоте Красноярскому краю в 48 населенных пунктов.  Населенный пункт не включён в реализацию проекта на 2023 год. Информация о проведении онлайн-голосования по выбору населенных пунктов для реализации проекта в 2024 году будет опубликована на портале государственных услуг, а также на официальном сайте министерства цифрового развития Красноярского края по ссылке: http://www.digital.krskstate.ru/.]]]"


},

        {"role": "user", "content": trimmed_info},




    ]

    response = openai.ChatCompletion.create(
        model="gpt-4-turbo-preview",
        messages=messages,
        max_tokens=1500,
        n=1,
        temperature=0.8,
        stream=True,
    )
    message_content = ""  # Переменная для хранения состояния сообщения с разметкой
    message_content_no_md = ""  # Переменная для хранения состояния сообщения без разметки
    chunk_counter = 0  # Счетчик количества полученных фрагментов
    message = None  # Переменная для хранения объекта сообщения

    try:
        while True:
            chunk = next(response)  # Получаем следующий фрагмент

            if chunk["object"] == "error":  # Проверяем, является ли фрагмент ошибкой
                # обработка ошибки
                break

            delta = chunk.get("choices", [{}])[0].get("delta", {})  # Получаем дельту фрагмента
            message_delta = delta.get("content")  # Извлекаем текст дельты

            if message_delta is not None and message_delta.strip():  # Проверяем, не является ли текст пустым
                message_content += message_delta  # Добавляем содержимое дельты к текущему сообщению с разметкой
                message_content_no_md += remove_markdown(message_delta)  # Удаляем разметку и добавляем содержимое дельты к текущему сообщению без разметки
                chunk_counter += 1  # Увеличиваем счетчик фрагментов

            if chunk_counter % 30 == 0 or chunk["object"] == "chat.completion":  # Если набрано 20 фрагментов или это последний фрагмент
                if message_content_no_md:  # Проверяем, что содержимое без разметки не является пустой строкой
                    if message is None:  # Если сообщение еще не создано, создаем его
                        message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                        message = await bot.send_message(chat_id=chat_id, text=message_content_no_md)
                    else:  # Если сообщение уже было создано
                        if message.text != message_content_no_md:  # Проверяем, изменился ли текст сообщения
                            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                            try:
                                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content_no_md)
                            except aiogram.exceptions.MessageNotModified:
                                pass

            await asyncio.sleep(0.03)  # Добавляем небольшую задержку между фрагментами

    except StopIteration:
        pass

    if message_content and (message is None or message.text != message_content):
        if message is None:  # Если сообщение еще не создано, создаем его
            message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            message = await bot.send_message(chat_id=chat_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
        else:  # Если сообщение уже было создано
            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            try:
                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
            except aiogram.exceptions.MessageNotModified:
                pass

    if message_content and message is not None:
        return message.text

    return None

#10

import random
import threading
import time
import requests

async def handle_digital_ministry_info(call):
    global info_text_storage
    chat_id = call.message.chat.id

    sticker_ids = [
        "CAACAgIAAxkBAAEheFJkau4Sf6bNrjPQYlRkse5dpzz9FgACKT0AAulVBRhAF4Dz8yP3WS8E",
        "CAACAgIAAxkBAAEheFpkau8kZYI6B9gjs0HVsMgthiz9KgACBQEAAvcCyA_R5XS3RiWkoS8E",
        "CAACAgIAAxkBAAEheFhkau7n4LJwuPn7gPvW6ToYh9rjAgAC6BUAAiMlyUtQqGgG1fAXAAEvBA",
        "CAACAgIAAxkBAAEheF9kau9O_07D2mrWk4Oygs5DYHtaQgACgQEAAiteUwteCmw-bAABeLQvBA"
    ]

    random_sticker_id = random.choice(sticker_ids)

    sticker_message = await bot.send_sticker(chat_id, random_sticker_id)  # сохраняем сообщение со стикером

    info_text = info_text_storage[chat_id]

    async def send_typing(max_attempts=30):  # Максимальное количество попыток = 30
        attempts = 0
        while not message_sent and attempts < max_attempts:
            await bot.send_chat_action(chat_id, 'typing')
            time.sleep(5)  # Пауза между действиями "typing"
            attempts += 1

    message_sent = False
    threading.Thread(target=send_typing).start()

    try:
        openai_response = await send_request_to_openai(chat_id)
        #await bot.send_message(chat_id, openai_response)
    finally:
        message_sent = True
        await bot.delete_message(chat_id, sticker_message.message_id)  # Удаление сообщения со стикером после отправки основного сообщения


async def handle_digital_ministry_info_post(call):
    global info_text_storage
    chat_id = call.message.chat.id

    sticker_id = "CAACAgIAAxkBAAEhAotkX1-uvG-2lf3mufIMFqwDvpRyYwACKRUAAviLwEupQBIzh-Q46C8E"
    await bot.send_sticker(call.message.chat.id, sticker_id)

    info_text = info_text_storage[chat_id]

    async def send_typing(max_attempts=30):  # Максимальное количество попыток = 30
        attempts = 0
        while not message_sent and attempts < max_attempts:
            await bot.send_chat_action(chat_id, 'typing')
            time.sleep(5)  # Пауза между действиями "typing"
            attempts += 1

    message_sent = False
    threading.Thread(target=send_typing).start()

    try:
        openai_response = await send_request_to_openai_post(chat_id)
        await bot.send_message(chat_id, openai_response)
    except Exception as e:
        #print(f"Error while sending message: {e}")
        pass  # Добавьте "pass" после "except"

    finally:
        message_sent = True


weather_emoji_mapping = {
    'thunderstorm with light rain': '⛈️',
    'thunderstorm with rain': '⛈️',
    'thunderstorm with heavy rain': '⛈️',
    'light thunderstorm': '⛈️',
    'thunderstorm': '⛈️',
    'heavy thunderstorm': '⛈️',
    'ragged thunderstorm': '⛈️',
    'thunderstorm with light drizzle': '⛈️',
    'thunderstorm with drizzle': '⛈️',
    'thunderstorm with heavy drizzle': '⛈️',
    'light intensity drizzle': '🌧️',
    'drizzle': '🌧️',
    'heavy intensity drizzle': '🌧️',
    'light intensity drizzle rain': '🌧️',
    'drizzle rain': '🌧️',
    'heavy intensity drizzle rain': '🌧️',
    'shower rain and drizzle': '🌧️',
    'heavy shower rain and drizzle': '🌧️',
    'shower drizzle': '🌧️',
    'light rain': '🌧️',
    'moderate rain': '🌧️',
    'heavy intensity rain': '🌧️',
    'very heavy rain': '🌧️',
    'extreme rain': '🌧️',
    'freezing rain': '🌧️',
    'light intensity shower rain': '🌧️',
    'shower rain': '🌧️',
    'heavy intensity shower rain': '🌧️',
    'ragged shower rain': '🌧️',
    'light snow': '❄️',
    'snow': '❄️',
    'heavy snow': '❄️',
    'sleet': '❄️',
    'light shower sleet': '❄️',
    'shower sleet': '❄️',
    'light rain and snow': '❄️',
    'rain and snow': '❄️',
    'light shower snow': '❄️',
    'shower snow': '❄️',
    'heavy shower snow': '❄️',
    'mist': '🌫️',
    'smoke': '🌫️',
    'haze': '🌫️',
    'sand/dust whirls': '🌫️',
    'fog': '🌫️',
    'sand': '🌫️',
    'dust': '🌫️',
    'volcanic ash': '🌫️',
    'squalls': '🌫️',
    'tornado': '🌪️',
    'clear sky': '☀️',
    'few clouds: 11-25%': '🌤️',
    'scattered clouds: 25-50%': '🌥️',
    'broken clouds: 51-84%': '☁️',
    'overcast clouds: 85-100%': '☁️'
}


async def get_weather(latitude, longitude, api_key):
    url = f"http://api.openweathermap.org/data/2.5/weather?lat={latitude}&lon={longitude}&appid={api_key}&units=metric"
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            if response.status == 200:
                weather_data = await response.json()
                temp = weather_data["main"]["temp"]
                description = weather_data["weather"][0]["description"]
                emoji = weather_emoji_mapping.get(description, '')  # Получаем соответствующий эмодзи или пустую строку, если сопоставление не найдено
                return f"{temp}°C {emoji}"
            else:
                return "Не удалось получить информацию о погоде."



#11
async def animated_text(bot, chat_id, message_id, base_text, final_text, delay=0.003):
    current_text = base_text
    unchanged_count = 0
    for i in range(len(base_text), len(final_text), 2):  # измените здесь на 2
        await bot.send_chat_action(chat_id, 'typing')
        await asyncio.sleep(delay)
        # Если остался только один символ, добавьте его, иначе добавьте два символа
        current_text += final_text[i:i + 2] if i + 2 <= len(final_text) else final_text[i:i + 1]
        if current_text != base_text:
            try:
                await bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=current_text)
                base_text = current_text
                unchanged_count = 0
            except telebot.apihelper.ApiTelegramException as e:
                if 'message is not modified' in str(e):
                    unchanged_count += 1
                    if unchanged_count >= 3:
                        break
                else:
                    raise e
        await asyncio.sleep(delay)


def escape_markdown(text):
    markdown_escape_characters = ['*', '_', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    return re.sub('([{}])'.format(''.join(markdown_escape_characters)), r'\\\1', text)


is_main_menu_button_active = {}


def set_main_menu_button_active(chat_id, active):
    global is_main_menu_button_active
    is_main_menu_button_active[chat_id] = active


def create_go_main_menu_inline_button():
    inline_keyboard = types.InlineKeyboardMarkup()
    main_menu_button = types.InlineKeyboardButton("Выйти из справочника", callback_data="go_main_menu")
    inline_keyboard.add(main_menu_button)
    return inline_keyboard


@dp.callback_query_handler(lambda call: call.data == "go_main_menu")
async def go_main_menu_callback_handler(call):
    set_main_menu_button_active(call.message.chat.id, False)
    await bot.send_message(call.message.chat.id,
                           "Вы вышли из справочника\n\nВведите название населенного пункта, чтобы получить информацию о связи\nЧтобы узнать, кто сегодня в отпуске, жми /otpusk\nСправочник: /employee \nЕсли остались вопросы, пиши @rejoller.")
    bot.clear_step_handler_by_chat_id(call.message.chat.id)
    remove_employee_data(call.message.chat.id)
    # Здесь вызвать функцию для отправки главного меню, например:
    # await send_main_menu(call.message)


@dp.callback_query_handler(lambda call: call.data == "go_main_menu")
async def process_go_main_menu_callback(call):
    await go_main_menu_callback_handler(call)



def remove_employee_data(chat_id):
    if chat_id in stored_employees_data:
        del stored_employees_data[chat_id]

from aiogram.dispatcher.filters import Text


@dp.message_handler(Text(equals='Выйти из справочника'))
async def handle_go_main_menu(message: types.Message):
    set_main_menu_button_active(message.chat.id, False)
    await bot.send_message(message.chat.id,
                           'Вы вышли из справочника\n\nВведите название населенного пункта, чтобы получить информацию о связи\nЧтобы узнать, кто сегодня в отпуске, жми /otpusk\nСправочник: /employee \nЕсли остались вопросы, пиши @rejoller.')
    bot.clear_step_handler(message)
    remove_employee_data(message.chat.id)
    # Здесь вызвать функцию для отправки главного меню, например:
    # await send_main_menu(message)





async def handle_employee_info_message(chat_id, response):
    set_main_menu_button_active(chat_id, True)
    main_menu_inline_button = create_go_main_menu_inline_button()
    msg = await bot.send_message(chat_id, response, parse_mode='HTML', reply_markup=main_menu_inline_button)
    bot.register_next_step_handler(msg, process_employee_name_input)  # добавим вызов функции process_employee_name_input здесь


@dp.message_handler(Text(equals='Выйти из справочника'))
async def process_go_main_menu_callback(message: types.Message):
    if message.text == 'Выйти из справочника':
        await go_main_menu_callback_handler(message)
    else:
        await bot.send_message(message.chat.id, "Введите фамилию сотрудника (и, при необходимости, имя и отчество)")
        dp.register_message_handler(process_employee_name_input, content_types=types.ContentTypes.TEXT, state="*")


import json



stored_employees_data = {}


def search_employee(name_parts):
    try:
        service = get_service()
        result = service.spreadsheets().values().get(
            spreadsheetId=SPREADSHEET_ID_1,
            range="Справочник!A2:J",
        ).execute()
        rows = result.get('values', [])
    except HttpError as error:
        # print(f"An error occurred: {error}")
        rows = []

    found_employees = []
    for row in rows:
        if row and len(row) > 2 and row[2]:
            fio_parts = re.split(r'\s+', row[2].strip())
            if len(fio_parts) >= len(name_parts) and all(p1.lower() == p2.lower() for p1, p2 in zip(fio_parts, name_parts)):
                found_employees.append(row)

    return found_employees


@dp.message_handler(commands=['employee'])
async def handle_employee_command(message: types.Message):
    global stored_employees_data
    user_first_name = message.from_user.first_name
    name_parts = re.split(r'\s+', message.text[len('/employee'):].strip())

    if len(name_parts) < 1 or len(name_parts[0]) < 2:
        msg = await message.reply(f'{user_first_name}, введите фамилию сотрудника (и, при необходимости, имя и отчество).')
        await msg.register_next_step_handler(process_employee_name_input)
        return

    found_employees = search_employee(name_parts)
    stored_employees_data[message.chat.id] = found_employees
    await process_searched_employee_results(message, found_employees)


async def process_employee_name_input(message: types.Message):
    global stored_employees_data
    name_parts = re.split(r'\s+', message.text.strip())

    if len(name_parts) < 1 or len(name_parts[0]) < 2:
        msg = await message.reply("Введите фамилию сотрудника (и, при необходимости, имя и отчество)")
        await msg.register_next_step_handler(process_employee_name_input)
        return

    found_employees = search_employee(name_parts)
    stored_employees_data[message.chat.id] = found_employees
    await process_searched_employee_results(message, found_employees)


def format_employee_info(employee):
    fio = escape_markdown(employee[2]) if len(employee) > 2 and employee[2] else "-"
    position = escape_markdown(employee[1]) if len(employee) > 1 and employee[1] else "-"
    department = escape_markdown(employee[0]) if len(employee) > 0 and employee[0] else "-"
    place = escape_markdown(employee[5]) if len(employee) > 5 and employee[5] else "-"
    workphone = escape_markdown(employee[3]) if len(employee) > 3 and employee[3] else "-"
    private_phone = escape_markdown(employee[4]) if len(employee) > 4 and employee[4] else "-"
    bd = escape_markdown(employee[7]) if len(employee) > 7 and employee[7] else "-"
    email = escape_markdown(employee[8]) if len(employee) > 8 and employee[8] else "-"
    tg = escape_markdown(employee[9]) if len(employee) > 9 and employee[9] else "-"

    response = (
        f"<b>ФИО:</b> {fio}\n"
        f"<b>Должность:</b> {position}\n"
        f"<b>Отдел:</b> {department}\n"
        f"<b>Место работы:</b> {place}\n"
        f"<b>Рабочий телефон:</b> {workphone}\n"
        f"<b>Сотовый телефон:</b> {private_phone}\n"
        f"<b>ДР:</b> {bd}\n"
        f"<b>Email:</b> {email}\n"
        f"<b>tg:</b> {tg}"
    )
    return response


async def process_searched_employee_results(message: types.Message, found_employees):
    if not found_employees:
        error_response = 'Не удалось найти информацию по данному запросу'
        await handle_employee_info_message(message.chat.id, error_response)
        return

    if len(found_employees) == 1:
        response = format_employee_info(found_employees[0])
        await handle_employee_info_message(message.chat.id, response)

    else:
        inline_keyboard = types.InlineKeyboardMarkup(row_width=1)

        for idx, employee in enumerate(found_employees):
            button_text = employee[2]  # ФИО из таблицы

            callback_data = json.dumps({"type": "employee_info", "index": idx, "chat_id": message.chat.id})
            button = types.InlineKeyboardButton(button_text, callback_data=callback_data)
            inline_keyboard.add(button)

        await bot.send_message(message.chat.id, "Выберите сотрудника из списка:", reply_markup=inline_keyboard)
        create_go_main_menu_inline_button()


@dp.callback_query_handler(lambda call: json.loads(call.data)["type"] == "employee_info")
async def handle_employee_info_call(call: types.CallbackQuery):
    global stored_employees_data
    data = json.loads(call.data)
    index = data["index"]
    chat_id = data["chat_id"]

    found_employees = stored_employees_data.get(chat_id)

    if found_employees and index < len(found_employees):
        employee = found_employees[index]
        response = format_employee_info(employee)
        await handle_employee_info_message(chat_id, response)
    else:
        await bot.send_message(chat_id, "Не удалось получить информацию о сотруднике.")
        send_go_main_menu_button(chat_id)

'''

async def send_request_to_gpt(chat_id, response):
    global info_text_storage
    trimmed_info = info_text_storage.get(chat_id, "")[:7000]
    messages = [
        {"role": "system", "content": "ты специалист по написанию обзоров на населенные пункты Красноярского края для телеграм бота Министерства цифрового развития - MCR info. пользователь будет тебе отправлять краткую информацию, а ты писать обзоры. Добавляй юмор, избегай темы политики и религии. Можешь добавлять цитаты философов, ученых, российских поэтов или цитаты из различных вселенных, таких как: MARVEL, гарри поттер, мортал комбат на свое усмотрение. Название населенного пункта пиши жирным шрифтом, цитаты курсивом. Разметка - Markdown. Будь деликатным, избегай высмеивания плохой связи, так как жители могут обидеться на тебя. Добавляй отступы, чтобы информация выглядела красиво. Не видоизменяй информацию в разделах - 'Информация из таблицы УЦН 2.0 (2023):' и 'Информация из таблицы 2023:'. Будь максимально лаконичен, не перегружай пользователя информацией! "},
        {"role": "user", "content": "Красноярск го, д. Песчанка (19.95°C )\n\n"
                                    "👥Население (2010 г.): 765 чел.\n"
                                    "👥Население (2020 г.): 762 чел.\n\n"
                                    "Сотовая связь: Билайн(3G 🟡Низкое)\n"
                                    "Мегафон(3G 🟡Низкое)\n"
                                    "МТС(3G 🟢Хорошее)\n"
                                    "Теле2(4G 🟡Низкое)\n"
                                    "Интернет: \n\n"
                                    "Количество таксофонов: 0\n"
                                    "Информация из таблицы УЦН 2.0 (2023):\n\n"
                                    "Голосов через Госусулуги: 4\n"
                                    "Голосов почтой: 0\n"
                                    "Итого голосов: 4\n"
                                    "Проверка услуг:\n"
                                    "Примечание: Субсидия 2017\n"
                                    "Попадение в квоту:\n\n"
                                    "Чтобы узнать кто сегодня в отпуске жми /otpusk\n"
                                    "Если нужен справочник, жми /employee"},
        {"role": "assistant", "content": "🏞 Красноярск го, д. Песчанка (25.9)\n)"
                                         "👥 Население (2010 г.): 765 чел.\n"
                                         "🌍 Население (2020 г.): 762 чел.\n\n"
                                         "💬 Я хотел бы рассказать тебе о Песчанке, но увы, информация о ней так же редка, как песчинки в пустыне. Но не отчаивайся, ведь у нас есть другие интересные факты о населенных пунктах Красноярского края! 😄\n"
                                         "📶 Уровень связи:\n"
                                         "Сотовая связь:\n"
                                         "- Билайн: 3G 🟡 Низкое (Можно отправить СМС, но с фотками лучше подождать)\n"
                                         "- Мегафон: 3G 🟡 Низкое (Позвонить можно, но видеозвонок может вызвать нервный тик)\n"
                                         "- МТС: 3G 🟢 Хорошее (Связь стабильна, как самый преданный друг)\n"
                                         "- Теле2: 4G 🟡 Низкое (Скорость, как у черепахи, но лучше, чем ничего)\n\n"
                                         "🌐 Интернет: \n"
                                         "- Wi-Fi: Прогулка по соседним улицам может привести к счастливому случаю - пойманному Wi-Fi сигналу от какого-нибудь кафе или магазина. 📶🔍\n"
                                         "- Домашний интернет: Возможно, лучше отложить онлайн-гейминг на другой раз, но для чтения новостей и просмотра видео хватит. 🏡💻 \n\n"
                                         " Таксофоны: 0\n\n"
                                         "Информация из таблицы УЦН 2.0 (2023):\n\n"
                                         "Голосов через Госусулуги: 4\n"
                                         "Голосов почтой: 0\n"
                                         "Итого голосов: 4\n"
                                         "Проверка услуг:\n"
                                         "Примечание: Субсидия 2017\n"
                                         "Попадение в квоту:\n\n"
                                         "🕵️‍♀️ Если тебе нужна информация о сотрудниках или кто сегодня в отпуске, просто нажми /employee или /otpusk.\n\n"
                                         "🏝️ Пока что это все, что я знаю о Песчанке. Но не сдавайся, в следующий раз я приведу еще больше интересной информации! 💪😉"},
        {"role": "user", "content": "Отлично, а теперь напиши подобным образом для этого населенного пункта: "},
        {"role": "user", "content": response}
    ]

    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo-16k-0613",
        messages=messages,
        max_tokens=2000,
        n=1,
        temperature=1,
        stream=True,
    )
    message_content = ""  # Переменная для хранения состояния сообщения
    chunk_counter = 0  # Счетчик количества полученных фрагментов
    message = None  # Переменная для хранения объекта сообщения

    try:
        while True:
            chunk = response['choices'][0]['message']['content']  # Получаем следующий фрагмент

            print("Received chunk:")
            print(chunk)

            if "error" in chunk:  # Проверяем, является ли фрагмент ошибкой
                # обработка ошибки
                break

            delta = response['choices'][0]['message'].get("delta", {})  # Получаем дельту фрагмента
            message_delta = delta.get("content")  # Извлекаем текст дельты

            if message_delta is not None and message_delta.strip():  # Проверяем, не является ли текст пустым
                message_content += message_delta  # Добавляем содержимое дельты к текущему сообщению
                chunk_counter += 1  # Увеличиваем счетчик фрагментов

            if chunk_counter % 20 == 0:  # Если набрано 20 фрагментов
                if message_content:  # Проверяем, что содержимое не является пустой строкой
                    if message is None:  # Если сообщение еще не создано, создаем его
                        message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                        message = await bot.send_message(chat_id=chat_id, text=escape(message_content), parse_mode=types.ParseMode.HTML)
                    else:  # Если сообщение уже было создано
                        if message.text != message_content:  # Проверяем, изменился ли текст сообщения
                            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                            message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=escape(message_content), parse_mode=types.ParseMode.HTML)
                    print("Current message:")
                    print(message_content)
                else:
                    print("Empty message content, skipping...")

            await asyncio.sleep(0.5)  # Добавляем небольшую задержку между фрагментами

    except StopIteration:
        pass

    if message_content and message is not None:
        return message.text
'''

async def send_request_to_gpt(chat_id, response):
    global info_text_storage
    trimmed_info = info_text_storage.get(chat_id, "")[:7000]
    messages = [
        {"role": "system", "content": "ты специалист по написанию обзоров на населенные пункты Красноярского края для телеграм бота Министерства цифрового развития - MCR info. пользователь будет тебе отправлять краткую информацию, а ты писать обзоры. Тон нейтральный, избегай темы политики и религии. Можешь добавлять цитаты ученых и русских писателей. Будь деликатным, избегай высмеивания плохой связи, так как жители могут обидеться на тебя. Добавляй отступы, чтобы информация выглядела красиво. Не видоизменяй информацию в разделах - 'Информация из таблицы УЦН 2.0 (2023):' и 'Информация из таблицы 2023:'. Будь максимально лаконичен, не перегружай пользователя информацией! Не забывай пропускать строки с помощью [\n]. Разметка markdown."},
        {"role": "user", "content": "Красноярск го, д. Песчанка (19.95°C )\n\n"
                                    "👥Население (2010 г.): 765 чел.\n"
                                    "👥Население (2020 г.): 762 чел.\n\n"
                                    "Сотовая связь: Билайн(3G 🟡Низкое)\n"
                                    "Мегафон(3G 🟡Низкое)\n"
                                    "МТС(3G 🟢Хорошее)\n"
                                    "Теле2(4G 🟡Низкое)\n"
                                    "Интернет: \n\n"
                                    "Количество таксофонов: 0\n"
                                    "Информация из таблицы УЦН 2.0 (2023):\n\n"
                                    "Голосов через Госусулуги: 4\n"
                                    "Голосов почтой: 0\n"
                                    "Итого голосов: 4\n"

                                    "Проверка услуг:\n"
                                    "Примечание: Субсидия 2017\n"
                                    "Попадение в квоту:\n\n"
                                    "Чтобы узнать кто сегодня в отпуске жми /otpusk\n"
                                    "Если нужен справочник, жми /employee"},
        {"role": "assistant", "content": "🏞 *Красноярск го, д. Песчанка (25.9°C)*\n\n\n"
                                         "\n👥 Население (2010 г.): 765 чел.\n"
                                         "\n🌍 Население (2020 г.): 762 чел.\n\n"
                                         "💬 Я хотел бы рассказать тебе о Песчанке, но увы, информация о ней так же редка, как песчинки в пустыне. Но не отчаивайся, ведь у нас есть другие интересные факты о населенных пунктах Красноярского края! 😄\n"
                                         "📶 Уровень связи:\n"
                                         "Сотовая связь:\n"
                                         "- Билайн: 3G 🟡 Низкое \n"
                                         "- Мегафон: 3G 🟡 Низкое \n"
                                         "- МТС: 3G 🟢 Хорошее \n"
                                         "- Теле2: 4G 🟡 Низкое \n"
                                         "В целом качество сотовой связи в этом населенном пункте ниже среднего!\n\n"
                                         "🌐 Интернет: \n"

                                         "- Информации о домашнем интернете у меня нет, но в населенном пункте есть мобильный интернет. Возможно, лучше отложить онлайн-гейминг на другой раз, но для чтения новостей и просмотра видео хватит. 🏡💻 \n\n"
                                         "☎️Таксофоны: 0 \n"

                                         "\nНо вместо таксофонов ты можешь пользоваться сотовой связью и интернетом\n\n"
                                         "\nИнформация из таблицы УЦН 2.0 (2023):\n\n"
                                         "Голосов через Госусулуги: 4\n"
                                         "Голосов почтой: 0\n"
                                         "Итого голосов: 4\n"
                                         "Проверка услуг:\n"
                                         "Примечание: Субсидия 2017\n"
                                         "Попадение в квоту:\n\n"

                                         "🕵️‍♀️ Если тебе нужна информация о сотрудниках или кто сегодня в отпуске, просто нажми /employee или /otpusk.\n\n"
                                         "🏝️ Пока что это все, что я знаю о Песчанке. Если тебе нужна информация о других населенных пунктах края, просто напиши мне и я расскажу тебе о качестве связи и о программах которые в нем реализуются!💪😉"},
        {"role": "user", "content": "Отлично, а теперь напиши подобным образом для этого населенного пункта: "},
        {"role": "user", "content": response}
    ]

    response = openai.ChatCompletion.create(
        model="gpt-4-turbo-preview",
        messages=messages,
        max_tokens=2000,
        n=1,
        temperature=1,
        stream=True,
    )

    message_content = ""  # Переменная для хранения состояния сообщения с разметкой
    message_content_no_md = ""  # Переменная для хранения состояния сообщения без разметки
    chunk_counter = 0  # Счетчик количества полученных фрагментов
    message = None  # Переменная для хранения объекта сообщения

    try:
        while True:
            chunk = next(response)  # Получаем следующий фрагмент

            if chunk["object"] == "error":  # Проверяем, является ли фрагмент ошибкой
                # обработка ошибки
                break

            delta = chunk.get("choices", [{}])[0].get("delta", {})  # Получаем дельту фрагмента
            message_delta = delta.get("content")  # Извлекаем текст дельты

            if message_delta is not None and message_delta.strip():  # Проверяем, не является ли текст пустым
                print(f"Adding to message_content: {message_delta}")
                message_content += message_delta  # Добавляем содержимое дельты к текущему сообщению с разметкой
                message_content_no_md += remove_markdown(message_delta)  # Удаляем разметку и добавляем содержимое дельты к текущему сообщению без разметки
                chunk_counter += 1  # Увеличиваем счетчик фрагментов




            if chunk_counter % 20 == 0 or chunk["object"] == "chat.completion":  # Если набрано 20 фрагментов или это последний фрагмент
                if message_content_no_md:  # Проверяем, что содержимое без разметки не является пустой строкой
                    if message is None:  # Если сообщение еще не создано, создаем его
                        message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                        message = await bot.send_message(chat_id=chat_id, text=message_content_no_md)
                    else:  # Если сообщение уже было создано
                        if message.text != message_content_no_md:  # Проверяем, изменился ли текст сообщения
                            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                            try:
                                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content_no_md)
                            except aiogram.exceptions.MessageNotModified:
                                pass

            await asyncio.sleep(0.05)  # Добавляем небольшую задержку между фрагментами

    except StopIteration:
        pass


    if message_content and (message is None or message.text != message_content):
        if message is None:  # Если сообщение еще не создано, создаем его
            message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            message = await bot.send_message(chat_id=chat_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
        else:  # Если сообщение уже было создано
            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            try:
                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
            except aiogram.exceptions.MessageNotModified:
                pass

    if message_content and message is not None:
        return message.text

    return None
'''
async def send_request_to_gpt(chat_id, response):
    global info_text_storage
    trimmed_info = info_text_storage.get(chat_id, "")[:7000]
    messages = [
        {"role": "system", "content": "Ты виртуальный помощник MCR_info. Просто повторяй то, что тебе присылает пользователь, процент правок - до 7%. Используй русский язык."},

        {"role": "user", "content": response}
    ]

    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo-16k-0613",
        messages=messages,
        max_tokens=1000,
        n=1,
        temperature=1,
        stream=True,
    )

    message_content = ""  # Переменная для хранения состояния сообщения
    chunk_counter = 0  # Счетчик количества полученных фрагментов
    message = None  # Переменная для хранения объекта сообщения

    try:
        while True:
            chunk = next(response)  # Получаем следующий фрагмент

           # print("Received chunk:")
           # print(chunk)

            if chunk["object"] == "error":  # Проверяем, является ли фрагмент ошибкой
                # обработка ошибки
                break

            delta = chunk.get("choices", [{}])[0].get("delta", {})  # Получаем дельту фрагмента
            message_delta = delta.get("content")  # Извлекаем текст дельты

            if message_delta is not None and message_delta.strip():  # Проверяем, не является ли текст пустым
                message_content += message_delta  # Добавляем содержимое дельты к текущему сообщению
                chunk_counter += 1  # Увеличиваем счетчик фрагментов

            if chunk_counter % 20 == 0 or chunk["object"] == "chat.completion":  # Если набрано 20 фрагментов или это последний фрагмент
                if message_content:  # Проверяем, что содержимое не является пустой строкой
                    if message is None:  # Если сообщение еще не создано, создаем его
                        message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                        message = await bot.send_message(chat_id=chat_id, text=message_content)
                    else:  # Если сообщение уже было создано
                        if message.text != message_content:  # Проверяем, изменился ли текст сообщения
                            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
                            try:
                                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content)
                            except aiogram.exceptions.MessageNotModified:
                                pass

                    #print("Current message:")
                   # print(message_content)
                else:
                    print("Empty message content, skipping...")

            await asyncio.sleep(0.03)  # Добавляем небольшую задержку между фрагментами

    except StopIteration:
        pass

    if message_content and (message is None or message.text != message_content):
        if message is None:  # Если сообщение еще не создано, создаем его
            message = await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            message = await bot.send_message(chat_id=chat_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
        else:  # Если сообщение уже было создано
            await bot.send_chat_action(chat_id=chat_id, action=types.ChatActions.TYPING)  # Отправляем статус "typing"
            try:
                message = await bot.edit_message_text(chat_id=chat_id, message_id=message.message_id, text=message_content, parse_mode=types.ParseMode.MARKDOWN)
            except aiogram.exceptions.MessageNotModified:
                pass



    if message_content and message is not None:
        return message.text

    return None

'''


#13
async def handle_additional_info(query):
    chat_id = json.loads(query.data)["chat_id"]
    if chat_id in additional_info_storage:
        response = additional_info_storage[chat_id]
        messages = split_message(response)
        for message_group in messages:
            msg = ''.join(message_group)
            if msg.strip():  # Проверка, что сообщение не пустое
                await bot.send_message(chat_id, msg, parse_mode='Markdown')

        await bot.answer_callback_query(query.id)
    else:
        await bot.answer_callback_query(query.id, "Дополнительная информация недоступна.")


async def handle_espd_info(query):
    chat_id = json.loads(query.data)["chat_id"]
    if chat_id in espd_info_storage:
        response = espd_info_storage[chat_id]
        messages = split_message(response)
        for message_group in messages:
            msg = ''.join(message_group)
            if msg.strip():  # Проверка, что сообщение не пустое
                #msg = msg.replace('*', '\\*').replace('_', '\\_').replace('[', '\\[').replace(']', '\\]')

                await bot.send_message(chat_id, msg, parse_mode='HTML')

        await bot.answer_callback_query(query.id)
    else:
        await bot.answer_callback_query(query.id, "Информация из таблицы ЕСПД недоступна.")


async def handle_szoreg_info(query):
    chat_id = json.loads(query.data)["chat_id"]
    if chat_id in szoreg_info_storage:
        response = szoreg_info_storage[chat_id]
        messages = split_message(response)
        for message_group in messages:
            msg = ''.join(message_group)
            if msg.strip():  # Проверка, что сообщение не пустое
                await bot.send_message(chat_id, msg, parse_mode='HTML')

        await bot.answer_callback_query(query.id)
    else:
        await bot.answer_callback_query(query.id, "Информация из таблицы СЗО (региональный контракт) недоступна.")


async def handle_schools_info(query):
    chat_id = json.loads(query.data)["chat_id"]
    if chat_id in schools_info_storage:
        response = schools_info_storage[chat_id]
        messages = split_message(response)
        for message_group in messages:
            msg = ''.join(message_group)
            if msg.strip():  # Проверка, что сообщение не пустое
                await bot.send_message(chat_id, msg, parse_mode='HTML')

        await bot.answer_callback_query(query.id)
    else:
        await bot.answer_callback_query(query.id, "Информация из таблицы по школам недоступна.")



from itertools import chain

async def handle_survey_chart(query):
    print("handle_survey_chart called with query data:", query.data)
    chat_id = json.loads(query.data)["chat_id"]

    if chat_id in survey_data_storage:
        logging.info("Data found for chat_id %s", chat_id)
        survey_data = survey_data_storage[chat_id]

        data_list = []

        # Создание списка словарей с данными для каждой строки
        for data_row in survey_data:
            # Убедимся, что у нас есть достаточно данных для каждого столбца
            data_row += [None] * (14 - len(data_row))

            row_data = {
                "Дата": data_row[0],
                "ID": data_row[1],
                "Имя": data_row[2],
                "Фамилия": data_row[3],
                "Ник": data_row[4],
                "Ключ": data_row[5],
                "Уровень_Tele2": data_row[6],
                "Качество_Tele2": data_row[7],
                "Уровень_Megafon": data_row[8],
                "Качество_Megafon": data_row[9],
                "Уровень_Beeline": data_row[10],
                "Качество_Beeline": data_row[11],
                "Уровень_MTS": data_row[12],
                "Качество_MTS": data_row[13],
            }
            data_list.append(row_data)

        # Преобразование данных в DataFrame
        data_df = pd.DataFrame(data_list)
        print("DataFrame created with data:", data_df)

        title = "Результаты опроса"  # Установите нужный заголовок для графика

        # Перебираем все строки в DataFrame
                # Перебираем все строки в DataFrame
      #  for idx, row in data_df.iterrows():
        try:
            #print(f"Calling create_individual_radar_chart for row {idx}...")
            await create_individual_radar_chart(chat_id, data_df, title)  # передаем весь DataFrame, а не одну строку
        except Exception as e:
            print("An error occurred:", e)
            logging.error("An error occurred: %s", e, exc_info=True)

    else:
        print(f"No data found for chat_id {chat_id}")












#@dp.message_handler(Command('text'))
dp.register_callback_query_handler(handle_additional_info, lambda query: json.loads(query.data)["type"] == "additional_info")
dp.register_callback_query_handler(handle_espd_info, lambda query: json.loads(query.data)["type"] == "espd_info")
dp.register_callback_query_handler(handle_szoreg_info, lambda query: json.loads(query.data)["type"] == "szoreg_info")
dp.register_callback_query_handler(handle_schools_info, lambda query: json.loads(query.data)["type"] == "schools_info")
dp.register_callback_query_handler(handle_digital_ministry_info, lambda query: json.loads(query.data)["type"] == "digital_ministry_info")
dp.register_callback_query_handler(handle_digital_ministry_info_post, lambda query: json.loads(query.data)["type"] == "digital_ministry_info_post")
dp.register_callback_query_handler(handle_survey_chart, lambda query: json.loads(query.data)["type"] == "survey_chart")







import unicodedata
from google.cloud import vision_v1 as vision
from pdf2image import convert_from_path
from docx import Document
from docx.shared import Pt, Cm
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT, WD_LINE_SPACING

from PIL import Image
import io
from google.oauth2 import service_account


def send_request_to_vision(text):
    messages = [
        {"role": "system", "content": "ты создан для того чтобы распознавать файлы для их вставки в документ ворд. В основном ты распознаешь отсканированную информацию с официальных писем. распознавай как самый превосходный сканер на планете. не вставляй лишние переносы строк, так как это портит вид итогового документа. Учти что в документе выравнивание текста должно быть по ширине и текст должен выглядеть красиво."},
        {"role": "user", "content": text},
    ]
    response = openai.ChatCompletion.create(
        model="gpt-4-turbo-preview",
        messages=messages,
        max_tokens=1500,
        n=1,
        temperature=1,
    )
    return response['choices'][0]['message']['content']

@dp.message_handler(content_types=types.ContentType.DOCUMENT)
async def handle_docs(message: types.Message):
    doc_id = message.document.file_id
    file_info = await bot.get_file(doc_id)
    downloaded_file = await bot.download_file(file_info.file_path)

    # Создание директории, если она не существует
    dir_path = '/home/rejoller/mcrbot/'
    os.makedirs(dir_path, exist_ok=True)

    # Сохранение файла на диск
    file_path = os.path.join(dir_path, 'file.pdf')
    with open(file_path, 'wb') as f:
        f.write(downloaded_file.read())

    # Конвертация PDF в список изображений
    images = convert_from_path(file_path)

    # Создание нового Word документа
    doc = Document()
    creds = service_account.Credentials.from_service_account_file(
            '/home/rejoller/mcrbot/credentials.json')

    client = vision.ImageAnnotatorClient(credentials=creds)
    full_text = ""
    for image in images:
        # Convert PIL Image to Bytes
        byte_arr = io.BytesIO()
        image.save(byte_arr, format='PNG')
        byte_arr = byte_arr.getvalue()
        image = vision.Image(content=byte_arr)

        # Делаем запрос к Google Cloud Vision API
        response = client.text_detection(image=image)
        texts = response.text_annotations

        # Сортируем блоки текста по их вертикальному положению на странице
        texts.sort(key=lambda text: text.bounding_poly.vertices[0].y)

        # Добавляем каждый блок текста в общий текст
        for text in texts:
            normalized_text = unicodedata.normalize('NFKD', text.description)
            cleaned_text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', normalized_text)
            full_text += " " + cleaned_text

    # Дополнительная нормализация текста с помощью GPT-3.5-turbo
    gpt_normalized_text = send_request_to_vision(full_text)

    paragraph = doc.add_paragraph(gpt_normalized_text)

    # Установка стиля и форматирования
    run = paragraph.runs[0]
    run.font.name = 'Times New Roman'
    run.font.size = Pt(14)
    paragraph_format = paragraph.paragraph_format
    paragraph_format.alignment = WD_PARAGRAPH_ALIGNMENT.JUSTIFY
    paragraph_format.line_spacing = 1
    paragraph_format.first_line_indent = Cm(1.25)

    # Save the Word document to a temporary file
    doc_path = os.path.join(dir_path, 'temp.docx')
    doc.save(doc_path)

    # Send the Word document back to the user
    with open(doc_path, "rb") as doc_file:
        await bot.send_document(message.chat.id, doc_file)







import difflib


import asyncio


















#@dp.message_handler(Command('text'))
@dp.message_handler()


async def handle_text(message: types.Message, state: FSMContext):
    user_state = await state.get_state()
    if user_state == Form.waiting_for_number.state:
        return



    print ('основная работает')
    global info_text_storage
    user_first_name = message.from_user.first_name
    await log_user_data_from_message(message)
    chat_id = message.chat.id



    user_id = message.from_user.id  # Получаем user_id
   # if user_id == 964635576:
    #    await bot.send_message(message.chat.id, "Для того чтобы пользоваться ботом отправьте сообщение муниципалам")
      #  return  # Завершаем выполнение функции
    votes_response = ""
    response = ''
    ucn2_response = ""
    operators_response = ''
    survey_results_values = ''
    '''
    if message.text.lower().startswith("обращение"):
        # Подготовьте и отправьте запрос в OpenAI
        await bot.send_message(message.chat.id, f'😈')
        response = await send_request_to_openai_2(message.chat.id, message.text[10:].strip())  # Извлечь информацию из сообщения после "обращение "
        await bot.send_message(message.chat.id, response)
        return

    if message.text.lower().startswith("тор "):
        await handle_tor_message(message)
        return
    if message.text.lower() == "кнд край":
        await handle_knd_kraj_message(message)
        return
    if message.text.lower().startswith("мсзу "):
        await bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌')
        await handle_mszu_message(message)
        return
    if message.text.lower().startswith("2мсзу "):
        await bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌 Загружаю информацию по МСЗУ муниципальных образований ')
        await handle_2mszu_message(message)
        return
    '''
    base_text = f'С'
    final_text = f'🧐Секундочку, {user_first_name}'
    await bot.send_message(message.chat.id, f'🧐Секундочку, {user_first_name}')
    # Замеряем время выполнения функции search_values
    start_time = time.time()

    gc, spreadsheet = await get_authorized_client_and_spreadsheet()
    found_values_a = await search_values(message.text, spreadsheet)



   # found_values_a, found_values_s = await search_values(message.text)
   # found_mszu_values = await check_mszu_column_b(message.text)

    end_time = time.time()
    execution_time = end_time - start_time
    print("Время выполнения функции search_values: ", execution_time, "секунд")

    if not found_values_a:
        # Проверяем метод Левенштейна с 70% совпадениями
        levenshtein_matches = await search_values_levenshtein(message.text, spreadsheet, threshold=0.4, max_results=5)

        if levenshtein_matches:
            unique_matches = set(levenshtein_matches)  # Используем множество, чтобы убрать повторяющиеся значения
            first_match = list(unique_matches)  # Преобразуем обратно в список
            formatted_matches = "\n".join([f'`{match}`' for match in first_match])  # Создаем строки с обратными кавычками для каждого значения
            await bot.send_message(message.chat.id, f'Проверьте правильность написания и попробуйте еще раз. Возможно вы имели в виду:\n(нажмите на населенный пункт, чтобы скопировать)\n\n{formatted_matches}', parse_mode='Markdown')
        else:
            await bot.send_message(message.chat.id, 'Не удалось найти информацию по данному запросу.\nПроверьте, правильно ли введено название населенного пункта и попробуйте еще раз')
        return

    # Если соответствие найдено в столбце A
    allowed_users = {964635576, 1063749463, 374056328, 572346758, 434872315, 1045874687, 1063749463, 487922464, 371098269, 402748716}
    if found_values_a:
        found_values = found_values_a
        await state.update_data(found_values=found_values)
      #  await bot.send_message(chat_id="430334520", text="Света, ты получаешь минус один балл!")
       # await bot.send_message(chat_id="964635576", text="Света, ты получаешь минус один балл!")



        if len(found_values) == 1:
            latitude = found_values[0][7]  # Широта находится в столбце H таблицы goroda2.0
            longitude = found_values[0][8]



            weather_info = await get_weather(latitude, longitude, "7cc8daec601b8354e0bc6350592d6f98")
            yandex_2023_response = ''
            pokazatel_504p_lines = []
           # ucn2_values = await search_in_ucn2(found_values[0][4])
           # yandex_2023_values = await search_yandex_2023_values(found_values[0][4])
            #pokazatel_504p_values = await search_in_pokazatel_504p(found_values[0][4])  # Используйте значение из столбца 4 в found_values_a



            if len(found_values) > 0 and len(found_values[0]) > 4:
                # Подразумевается, что если условие выполнено, то можно безопасно обращаться к found_values[5][4]

                ucn2_values, yandex_2023_values, pokazatel_504p_values, survey_results_values  = await asyncio.gather(
                    search_in_ucn2(found_values[0][4], spreadsheet),
                    search_yandex_2023_values(found_values[0][4], spreadsheet),
                    search_in_pokazatel_504p(found_values[0][4], spreadsheet),
                    search_in_results(found_values[0][4], spreadsheet)
                )
            else:
                # Если условие не выполнено, значит индекса [5][4] нет, и нужно обойтись без search_in_results
                ucn2_values, yandex_2023_values, pokazatel_504p_values = await asyncio.gather(
                    search_in_ucn2(found_values[0][4], spreadsheet),
                    search_yandex_2023_values(found_values[0][4], spreadsheet),
                    search_in_pokazatel_504p(found_values[0][4], spreadsheet)
                )
                survey_results_values = None



            '''
            try:
                votes_response = ''
                if len(found_values[0]) > 38:  # Убедитесь, что у вас достаточно данных в строке

                    votes = found_values[0][34] or "неизвестно"
                    update_time = found_values[0][35] or "неизвестно"
                    rank = found_values[0][36] or "неизвестно"
                    same_votes_np = found_values[0][38] or "неизвестно"

                    if votes != "неизвестно" and update_time != "неизвестно" and rank != "неизвестно" and same_votes_np != "неизвестно":
                        votes_response = f'\n\n<b>Голосование УЦН 2.0 2024</b>\n\n📊Количество голосов: <b>{votes}</b> (такое же количество голосов имеют {same_votes_np} населённых пунктов)\n🏆Место в рейтинге: {rank}\nДата обновления информации: {update_time}'
                    else:
                        print("Debug: Не все данные для блока голосования были найдены.")
                    response += votes_response
                else:
                    print("Debug: В строке данных недостаточно столбцов для извлечения информации о голосовании.")
            except Exception as e:
                print(f"Debug: Ошибка при извлечении данных о голосах: {e}")

            '''

            if found_values_a:
                for row in found_values_a:
                    # Создаем словарь с операторами и их значениями, используя метод get для безопасного обращения к элементам списка
                    operators = {
                        "Tele2": row[39] if len(row) > 39 else None,
                        "Мегафон": row[40] if len(row) > 40 else None,
                        "Билайн": row[41] if len(row) > 41 else None,
                        "МТС": row[42] if len(row) > 42 else None,
                    }
                    '''
                    operators_response = '\nОценка жителей:\n'
                    for operator_name, operator_value in operators.items():
                        if operator_value:  # Проверка на наличие значения (не None и не пустая строка)
                            # Переводим значение в строку, чтобы избежать ошибок при выполнении метода replace
                            operator_value = str(operator_value)
                            if "Отсутствует" in operator_value:
                                signal_level = "🔴Отсутствует"
                            elif "Низкое" in operator_value:
                                signal_level = "🟠Низкое"
                            elif "Среднее" in operator_value:
                                signal_level = "🟡Среднее"
                            elif "Хорошее" in operator_value:
                                signal_level = "🟢Хорошее"
                            else:
                                signal_level = "❓Неизвестно"

                            # Заменяем "(" и ")" на " "
                            signal_level = signal_level.replace("(", " ").replace(")", " ")
                            operators_response += f'{operator_name}: {signal_level}\n'

                    # Если ни один оператор не имеет данных, добавьте сообщение об отсутствии данных
                    if operators_response == '\nОценка жителей:\n':
                        operators_response += 'Нет данных\n'

                    response += operators_response
                    '''
                    operators_response = '\nОценка жителей:\n'

                    # Список для хранения ответов от операторов
                    operator_responses = []

                    for operator_name, operator_value in operators.items():

                        if operator_value:  # Проверка на наличие значения (не None и не пустая строка)
                            # Переводим значение в строку, чтобы избежать ошибок при выполнении метода replace
                            operator_value_str = str(operator_value)

                            # Пытаемся найти качество сигнала
                            signal_quality = re.search(r'Отсутствует|Низкое|Среднее|Хорошее', operator_value_str, re.IGNORECASE)
                            if signal_quality:
                                signal_quality = signal_quality.group()
                                if "Отсутствует" in signal_quality:
                                    signal_level = "🔴Отсутствует"
                                elif "Низкое" in signal_quality:
                                    signal_level = "🟠Низкое"
                                elif "Среднее" in signal_quality:
                                    signal_level = "🟡Среднее"
                                elif "Хорошее" in signal_quality:
                                    signal_level = "🟢Хорошее"
                                else:
                                    signal_level = "❓Неизвестно"

                                # Заменяем найденное качество сигнала на его эмодзи-эквивалент
                                operator_value_str = operator_value_str.replace(signal_quality, signal_level)
                            else:
                                operator_value_str = operator_value_str

                            # Заменяем "(" и ")" на " "
                            operator_value_str = operator_value_str.replace("(", " ").replace(")", " ")
                            operator_responses.append(f'{operator_name}: {operator_value_str}\n')
                        else:
                            # Если нет данных по оператору, не добавляем его в ответ
                            continue



                    # Если нет данных ни по одному оператору, добавляем сообщение об отсутствии данных
                    if not operator_responses:
                        operators_response += 'Нет данных\n'
                    else:
                        operators_response += ''.join(operator_responses)

                    response += operators_response




            if yandex_2023_values:
                yandex_2023_response = '\n\n\n<b>Информация из таблицы 2023</b>\n\n'
                for row in yandex_2023_values:
                    yandex_2023_response += f'Тип подключения: {row[4]}\nОператор: {row[15]}\nСоглашение: {row[7]}\nПодписание соглашения с МЦР: {row[8]}\nПодписание соглашения с АГЗ: {row[9]}\nДата подписания контракта: {row[11]}\nДата установки АМС: {row[12]}\nДата монтажа БС: {row[13]}\nЗапуск услуг: {row[14]}\n\n'
            if pokazatel_504p_values:
                for index in range(6, 10):
                    if len(pokazatel_504p_values[0]) > index and pokazatel_504p_values[0][index] and pokazatel_504p_values[0][index].strip():
                        value = pokazatel_504p_values[0][index]
                        if "Хорошее" in value:
                            value = value.replace("Хорошее", "🟢Хорошее")
                        if "Низкое" in value:
                            value = value.replace("Низкое", "🟠Низкое")
                        pokazatel_504p_lines.append(value)
            if ucn2_values:
                for row in ucn2_values:
                    ucn2_response = ''

                    if 4 < len(row) and row[4]:  # Проверка наличия значения
                        ucn2_response += '  Информация от Теле2:\n    -СМР: ' + row[4] + '\n'
                    if 5 < len(row) and row[5]:  # Проверка наличия значения
                        ucn2_response += '    -Запуск: ' + row[5] + '\n'
                    if 6 < len(row) and row[6]:  # Проверка наличия значения
                        ucn2_response += '    -Комментарии: ' + row[6] + '\n'

                    if ucn2_response:  # Если ucn2_response не пуст, добавить вводную строку в начало
                        ucn2_response = '\n\n\n<b>УЦН 2.0 2023</b>\n' + ucn2_response
                        response += ucn2_response







                response += ucn2_response

            pokazatel_504p_response = "\n".join(pokazatel_504p_lines) if pokazatel_504p_lines else "🔴отсутствует"


            if "4G" in pokazatel_504p_response:
                votes_response = ""
            else:
                # Теперь этот код выполняется только если "4G" не найдено в pokazatel_504p_response
                try:
                    if len(found_values[0]) > 38:  # Убедитесь, что у вас достаточно данных в строке
                        votes = found_values[0][34] or "неизвестно"
                        update_time = found_values[0][35] or "неизвестно"
                        rank = found_values[0][36] or "неизвестно"
                        same_votes_np = found_values[0][38] or "неизвестно"

                        if votes != "неизвестно" and update_time != "неизвестно" and rank != "неизвестно" and same_votes_np != "неизвестно":
                            votes_response = f'\n\n<b>Голосование УЦН 2.0 2024</b>\n\n📊Количество голосов: <b>{votes}</b> (такое же количество голосов имеют {same_votes_np} населённых пунктов)\n🏆Место в рейтинге: {rank}\nДата обновления информации: {update_time}'
                        else:
                            print("Debug: Не все данные для блока голосования были найдены.")
                except Exception as e:
                    print(f"Debug: Ошибка при извлечении данных о голосах: {e}")


          #  operators_response = await generate_operators_response([found_values_a])
            #print('operators_response', operators_response)  # Добавлено для отладки




          #  response = f'{found_values[0][1]}* ({weather_info})\n\n👥Население (2010 г.): {found_values[0][2]} чел.\n👥Население (2020 г.): {found_values[0][5]} чел.\n\nСотовая связь: {pokazatel_504p_response}\nИнтернет: {get_value(found_values[0], 9)}\n\nКоличество таксофонов: {get_value(found_values[0], 12)}{yandex_2023_response}{ucn2_response}\nЧтобы узнать кто сегодня в отпуске жми /otpusk \nЕсли нужен справочник, жми /employee'




          #  response = f'<b>{found_values[0][1]}</b> {weather_info}\n\n👥Население (2010 г.): {found_values[0][2]} чел.\n👥Население (2020 г.): {found_values[0][5]} чел.\n\n<b>Сотовая связь:</b>\n{pokazatel_504p_response}\n{operators_response}\n\nИнтернет: {await get_value(found_values[0], 9)}\n\nКоличество таксофонов: {await get_value(found_values[0], 12)}{yandex_2023_response}{ucn2_response}{votes_response}\n\nЕсли хочешь узнать о голосовании УЦН 2.0 2024 жми /votes\n\nБот для проведения опросов жителей - <a href="http://t.me/providers_rating_bot">@providers_rating_bot</a>'


            response = f'<b>{found_values[0][1]}</b>  {weather_info}'
            '''
            selsovet_info = await get_value(found_values[0], 20)
            tanya_sub_info_year = await get_value(found_values[0], 13)
            tanya_sub_info_provider = await get_value(found_values[0], 14)
            taksofony_info = await get_value(found_values[0], 12)
            arctic_info = await get_value(found_values[0], 6)
            internet_info = await get_value(found_values[0], 9)
            '''


            try:
                selsovet_info, tanya_sub_info_year, tanya_sub_info_provider, taksofony_info, arctic_info, internet_info, population_2010, population_2020, itog_ucn_2023 = await asyncio.gather(
                    get_value(found_values[0], 20),
                    get_value(found_values[0], 13),
                    get_value(found_values[0], 14),
                    get_value(found_values[0], 12),
                    get_value(found_values[0], 6),
                    get_value(found_values[0], 9),
                    get_value(found_values[0], 2),
                    get_value(found_values[0], 5),
                    get_value(found_values[0], 24),
                    return_exceptions=True  # Возврат исключений как объектов
                )
            except Exception as e:
                print(f"Произошла ошибка: {e}")


            if selsovet_info:
                response += f'\n{selsovet_info}'
            if arctic_info:
                response += f'\n❄️️арктическая зона❄️️'
            response += f'\n\n👥население 2010 г: {population_2010} чел.\n👥население 2020 г: {population_2020} чел.\n\n'

            if taksofony_info:
                response += f'\n☎️таксофон: {taksofony_info}'

            response += f'\n🌐интернет: {internet_info}️'
            response += f'\n\n📱<b>Сотовая связь:</b>\n{pokazatel_504p_response}'




            if tanya_sub_info_year and tanya_sub_info_provider:
                response += f'\n\nнаселенный пункт был подключен в рамках государственной программый "Развитие информационного общества" в {tanya_sub_info_year} году, оператор {tanya_sub_info_provider}'

            if itog_ucn_2023:
                response += f'\n\nкомментарий Минцифры России об УЦН 2024: {itog_ucn_2023}'
            response += f'\n{operators_response}\n'



            response += f'{yandex_2023_response}{ucn2_response}{votes_response}\nЕсли хочешь узнать о голосовании УЦН 2.0 2024 жми /votes\n\nБот для проведения опросов жителей - <a href="http://t.me/providers_rating_bot">@providers_rating_bot</a>'

            info_text_storage[message.chat.id] = response


            await bot.send_location(message.chat.id, latitude, longitude)
           # response = await send_request_to_gpt(chat_id, response)
           # await bot.send_message(chat_id="374056328", text="ку-ку")

            messages = split_message(response)



            for msg in messages:
                await bot.send_message(message.chat.id, msg, parse_mode='HTML', disable_web_page_preview=True)

        #    szofed_values = await search_szofed_values(found_values[0][4])
         #   espd_values = await search_espd_values(found_values[0][4])
         #   szoreg_values = await search_szoreg_values(found_values[0][4])
         #   schools_values = await search_schools_values(found_values[0][4])

          #  szofed_values, espd_values, szoreg_values, schools_values = await asyncio.gather(
           #     search_szofed_values(found_values[0][4], spreadsheet),
            #    search_espd_values(found_values[0][4], spreadsheet),
             #   search_szoreg_values(found_values[0][4], spreadsheet),
              #  search_schools_values(found_values[0][4], spreadsheet)
            #)


            szoreg_values, schools_values = await asyncio.gather(

                search_szoreg_values(found_values[0][4], spreadsheet),
                search_schools_values(found_values[0][4], spreadsheet)
            )
            print('szoreg:', szoreg_values)

            inline_keyboard = types.InlineKeyboardMarkup(row_width=3)

            if message.from_user.id in allowed_users:
                button_digital_ministry_info = types.InlineKeyboardButton("😈Подготовить ответ на обращение(БЕТА)", callback_data=json.dumps({"type": "digital_ministry_info", "chat_id": message.chat.id}))
                inline_keyboard.add(button_digital_ministry_info)
                #button_digital_ministry_info_post = types.InlineKeyboardButton("Сделать пост ВК", callback_data=json.dumps({"type": "digital_ministry_info_post", "chat_id": message.chat.id}))
                #inline_keyboard.add(button_digital_ministry_info_post)




            survey_data_storage[message.chat.id] = survey_results_values

            if survey_results_values:
                survey_inline_keyboard = types.InlineKeyboardMarkup()
                button_survey_results = types.InlineKeyboardButton("Показать результаты опроса", callback_data=json.dumps({"type": "survey_chart", "chat_id": message.chat.id}))
                survey_inline_keyboard.add(button_survey_results)
                await bot.send_message(message.chat.id, "Найдены результаты опроса. Хотите посмотреть?", reply_markup=survey_inline_keyboard)



          #  if szofed_values or espd_values or szoreg_values or schools_values or info_text_storage:
            if  szoreg_values or schools_values or info_text_storage:
                """
                if szofed_values:
                    szofed_response = '🏢*СЗО, подключенные в рамках федерального проекта в период с 2019 по 2021 год:*\n\n'
                    for i, row in enumerate(szofed_values, 1):
                        szofed_response += f'\n{i}. *Тип:* {row[8]}\n*Наименование:* {row[9]}\n*Адрес:* {row[4]}\n*Тип подключения (Узел связи):* {row[10]}\n*Пропускная способность:* {row[11]} Мб/сек\n*Дата подключения:* {row[12]}\n'

                    callback_data = json.dumps({"type": "additional_info", "chat_id": message.chat.id})
                    additional_info_storage[message.chat.id] = szofed_response
                    button_additional_info = types.InlineKeyboardButton("🏢СЗО", callback_data=callback_data)
                    inline_keyboard.add(button_additional_info)

                if espd_values:
                    espd_response = '🌐<b>Точки подключения к ЕСПД:</b> \n\n'
                    for i, row in enumerate(espd_values, 1):
                        espd_response += f'\n{i}. <b>Наименование:</b> {html.escape(row[12])}\n<b>Адрес:</b> {html.escape(row[8])}\n<b>Тип подключения:</b> {html.escape(row[9])}\n<b>Скорость:</b> {html.escape(row[13])}\n<b>Контакты:</b> {html.escape(row[18])}\n'

                    callback_data = json.dumps({"type": "espd_info", "chat_id": message.chat.id})
                    espd_info_storage[message.chat.id] = espd_response
                    button_espd_info = types.InlineKeyboardButton("🌐ЕСПД", callback_data=callback_data)
                    inline_keyboard.add(button_espd_info)
                """

                if szoreg_values:
                    szoreg_response = '🏢<b>Учреждения, подключенные по госпрограмме</b>\n\n'
                    for i, row in enumerate(szoreg_values, 1):

                        szoreg_response += f'\n{i}. <b>Тип:</b> {row[7]}\n<b>аименование:</b> {row[8]}\n<b>Адрес:</b> {row[5]} \n<b>Тип подключения:</b> {row[6]}\n<b>Пропускная способность:</b> {row[9]}\n<b>Контракт:</b> {row[10]}\n'

                        #if len(row) >= 10:

                         #   szoreg_response += f'<b>Комментарий:</b> {row[11]}\n'



                    callback_data = json.dumps({"type": "szoreg_info", "chat_id": message.chat.id})
                    szoreg_info_storage[message.chat.id] = szoreg_response
                    button_szoreg_info = types.InlineKeyboardButton(f"🏢Список учреждений ({len(szoreg_values)})",callback_data=callback_data)
                    inline_keyboard.add(button_szoreg_info)

                if schools_values:
                    schools_response = '🏫<b>Школы:</b>\n'
                    for i, row in enumerate(schools_values, 1):
                        schools_response += f'\n{i} '
                        if len(row) > 7:
                            schools_response += f'<b>{html.escape(row[12])}</b>\n'
                        if len(row) > 12:
                            schools_response += f'\n{html.escape(row[7])}\n'
                        if len(row) > 14:
                            schools_response += f'\n{html.escape(row[14])}, '
                        if len(row) > 13:
                            schools_response += f'{html.escape(row[13])} Мб/с\n'
                        if len(row) > 20:
                            schools_response += f'{html.escape(row[20])}'
                        schools_response += '\n'

                    callback_data = json.dumps({"type": "schools_info", "chat_id": message.chat.id})
                    schools_info_storage[message.chat.id] = schools_response
                    button_schools_info = types.InlineKeyboardButton("🏫Школы",callback_data=callback_data)
                    inline_keyboard.add(button_schools_info)

                await bot.send_message(message.chat.id, "⬇️Дополнительная информация⬇️", reply_markup=inline_keyboard)
            response_storage[message.chat.id] = response

            #dp.register_callback_query_handler(handle_additional_info, lambda query: json.loads(query.data)["type"] == "additional_info")
           # dp.register_callback_query_handler(handle_espd_info, lambda query: json.loads(query.data)["type"] == "espd_info")
          #  dp.register_callback_query_handler(handle_szoreg_info, lambda query: json.loads(query.data)["type"] == "szoreg_info")
         #   dp.register_callback_query_handler(handle_schools_info, lambda query: json.loads(query.data)["type"] == "schools_info")
        #    dp.register_callback_query_handler(handle_digital_ministry_info, lambda query: json.loads(query.data)["type"] == "digital_ministry_info")
       #     dp.register_callback_query_handler(handle_digital_ministry_info_post, lambda query: json.loads(query.data)["type"] == "digital_ministry_info_post")

        if len(found_values) > 1:
            values = [(await get_value(row, 1), await get_value(row, 2)) for row in found_values]
            values_with_numbers = [f"{i + 1}. {value[0]}" for i, value in enumerate(values)]
            msg = '\n'.join(values_with_numbers)

            messages = split_message(f'Найдено несколько населенных пунктов с таким названием. \n\n{msg}')

            for msg in messages:
                await bot.send_message(message.chat.id, msg)

            buttons = [str(i + 1) for i in range(len(found_values))]
            buttons.append("Отмена")
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)

            keyboard.add(*buttons)
            await bot.send_message(message.chat.id, 'Выберите номер необходимого населенного пункта:', reply_markup=keyboard)
            await Form.waiting_for_number.set()

    '''
    if found_values_s:
        found_values = found_values_s
        data = []
        for row in found_values_s:
            if len(row) >= 22:
                data.append([row[17], row[2], row[3], row[9], row[11], row[12], row[21]])
            else:
                print(f"Skipping row due to insufficient elements: {row}")

        if len(found_values) > 0:
            response = 'Наименование | Население | Сотовая связь | Интернет | Программа | Таксофон | СЗО (узел)\n'
            response += '-' * 71 + '\n'
            for row in found_values:
                if len(row) >= 22:
                    response += f"{row[17]} | {row[2]} | {row[3]} | {row[9]} | {row[11]} | {row[12]} | {row[19]}\n"

            excel_file = create_excel_file_2(headers, data)
            user_first_name = message.from_user.first_name

            file_name = found_values[0][18] if found_values else "table"

            with create_excel_file_2(headers, data) as file:
                file.name = f'{file_name}.xlsx'
                await bot.send_document(message.chat.id, file)

    if found_mszu_values:
        if len(found_mszu_values) == 1:
            response = f"\nНаименование услуги: \n{found_mszu_values[0][2]}\n№ в Рег.перечне (17-р): {found_mszu_values[0][0]}\nОтвет. РОИВ в рег.перечне: {found_mszu_values[0][3]}\nОтв. в плане: {found_mszu_values[0][6]}\nЕСНСИ: {found_mszu_values[0][8]}\nЕПГУ: {found_mszu_values[0][10]}"
            await bot.send_message(message.chat.id, response)

            inline_keyboard = types.InlineKeyboardMarkup()
            url_button = types.InlineKeyboardButton(text="Адрес ИФЗ", url=found_mszu_values[0][13])
            inline_keyboard.add(url_button)

            await bot.send_message(message.chat.id, "Нажмите на кнопку ниже для перехода по ссылке:", reply_markup=inline_keyboard)

        elif len(found_mszu_values) > 1:
            values_with_numbers = [f"{i + 1}. {value[2]}" for i, value in enumerate(found_mszu_values)]
            msg = '\n'.join(values_with_numbers)
            messages = split_message(f'Найдено несколько значений:\n\n{msg}')
            for msg in messages:
                await bot.send_message(message.chat.id, msg)

            buttons = [str(i + 1) for i in range(len(found_mszu_values))]
            buttons.append("Отмена")
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
            keyboard.add(*buttons)
            await bot.send_message(message.chat.id, 'Выберите номер необходимого значения:', reply_markup=keyboard)
            dp.register_message_handler(lambda x: handle_mszu_choice(x, found_mszu_valueskeyboard))

    '''
import random
from aiogram import types
from fuzzywuzzy import fuzz

async def search_espd_values_1(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('espd!A1:AL1466')
    rows = result.get('valueRanges', [])[0].get('values', [])

    found_values = []
    for row in rows:
        try:
            if len(row) > 12:
                location_column = row[6]
                name_column = row[12]
                functional_customer_column = row[11]

                location_ratio = fuzz.ratio(query.lower(), location_column.lower()) * 1.3  # Увеличиваем вес
                name_ratio = fuzz.ratio(query.lower(), name_column.lower())
                functional_customer_ratio = fuzz.ratio(query.lower(), functional_customer_column.lower())

                max_ratio = max(location_ratio, name_ratio, functional_customer_ratio)
                if max_ratio >= 20:
                    found_values.append((max_ratio, row))
        except Exception as e:
            print(f"Error processing row: {e}")

    sorted_values = sorted(found_values, key=lambda x: x[0], reverse=True)
    return [value[1] for value in sorted_values] or []





async def search_szoreg_values_1(query):
    agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
    gc = await agcm.authorize()
    spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
    result = await spreadsheet.values_batch_get('szoreg!A1:Q1700')
    rows = result.get('valueRanges', [])[0].get('values', [])

    found_values = []
    for row in rows:
        try:
            if len(row) > 8:
                location_column = row[3]
                type_institution_column = row[7]
                title_column = row[8]

                location_ratio = fuzz.ratio(query.lower(), location_column.lower()) * 1.3  # Увеличиваем вес
                type_institution_ratio = fuzz.ratio(query.lower(), type_institution_column.lower())
                title_ratio = fuzz.ratio(query.lower(), title_column.lower())

                max_ratio = max(location_ratio, type_institution_ratio, title_ratio)
                if max_ratio >= 20:
                    found_values.append((max_ratio, row))
        except Exception as e:
            print(f"Error processing row: {e}")

    sorted_values = sorted(found_values, key=lambda x: x[0], reverse=True)
    return [value[1] for value in sorted_values] or []






def get_thumb_url(column_value):
    thumb_url = None
    if column_value in ["минкульт", "МинКульт"]:
        thumb_url = "https://rejoller.pythonanywhere.com/static/minkult.png"
    elif column_value in ["ОИВ", "Аг.Занятости", "Аг.Печати", "ВетНадзор", "ГТН", "ЗАГС", "МинФин", "МинЛес", "МинОбр", "МинСоцПол", "МинСпорт", "МинФин", "МинСоцПол", "МинФин", "МирСуд", "МЭиРП", "МЦР", "СтройНадзор", "ЦИТ", "Ветнадзор", "МинСтрой", "ГТН"]:
        thumb_url = "https://rejoller.pythonanywhere.com/static/OIV.png"
    elif column_value in ["МинЗдрав", "ФАП"]:
        thumb_url = "https://rejoller.pythonanywhere.com/static/minzdrav.png"
    elif column_value in ["Аг.ГО и ЧС", "пожарная часть", "пож. часть", "пож. пост", "Аг.ГОиЧС"]:
        thumb_url = "https://rejoller.pythonanywhere.com/static/goichs.png"
    elif column_value == "ОМСУ":
        thumb_url = "https://rejoller.pythonanywhere.com/static/omsu.png"
    else:
        thumb_url = "https://rejoller.pythonanywhere.com/static/OIV.png"


    return thumb_url





@dp.inline_handler()
async def inline_query_handler(query: types.InlineQuery):
    if len(query.query) < 3:
        await query.answer(results=[], switch_pm_text="Результаты будут отображаться здесь", switch_pm_parameter="short_query")
        return
    found_values_a, _ = await search_values(query.query)
    found_espd_values = await search_espd_values_1(query.query)
    found_szoreg_values = await search_szoreg_values_1(query.query)

    thumb_urls_for_found_values_a = [
        "https://rejoller.pythonanywhere.com/static/town1.png",
        "https://rejoller.pythonanywhere.com/static/town2.png",
    ]
    results = []

    # Поиск в основных значениях
    for index, value_list in enumerate(found_values_a):
        location = f"<b>\n\n{value_list[1]}</b>" # Абанский р-н, п. Абан
        mobile_coverage_lines = []
        for coverage in value_list[29:33]:
            if coverage and coverage.strip(): # Проверяем, что строка не пуста
                if "Хорошее" in coverage:
                    coverage = f"🟢{coverage}"
                elif "Низкое" in coverage:
                    coverage = f"🟠{coverage}"
                mobile_coverage_lines.append(coverage)
        mobile_coverage = "\n".join(mobile_coverage_lines) if mobile_coverage_lines else "🔴отсутствует"
        population = int(value_list[5]) # население: 8207
        formatted_population = f"{population:,}".replace(',', ' ') # Добавляем пробелы в качестве разделителей разрядов
        value_str = f"{location}\n\nСотовая связь:\n{mobile_coverage}\nНаселение: {formatted_population}"
        thumb_url = random.choice(thumb_urls_for_found_values_a)
        bot_url = "https://t.me/MCRcoffee_bot" # URL чата с ботом
        result = types.InlineQueryResultArticle(
            id=str(index),
            title=value_list[1],
            input_message_content=types.InputTextMessageContent(message_text=value_str, parse_mode='HTML'),
            thumb_url=thumb_url,
            reply_markup=types.InlineKeyboardMarkup().add(
                types.InlineKeyboardButton("Перейти в чат с ботом", url=bot_url) # Кнопка ведет в чат с ботом
            )
        )
        results.append(result)

    espd_and_szoreg_results = []
    for row in found_espd_values:
        location_column = row[6]
        name_column = row[12]
        functional_customer_column = row[11]
        location_ratio = fuzz.ratio(query.query.lower(), location_column.lower()) * 1.3 # Исправлено здесь
        name_ratio = fuzz.ratio(query.query.lower(), name_column.lower()) # Исправлено здесь
        functional_customer_ratio = fuzz.ratio(query.query.lower(), functional_customer_column.lower()) # Исправлено здесь
        max_ratio = max(location_ratio, name_ratio, functional_customer_ratio)
        relevance = max_ratio
        functional_customer = row[11] # Функциональный заказчик
        address = row[8]
        additional_info = row[8]
        title = f"{functional_customer} - {additional_info}" # Комбинирование
        connection_type = row[9]
        speed = row[13]
        contacts = row[18]
        value_str = f'ЕСПД\n<b>Функциональный заказчик:</b> {functional_customer}\n<b>Наименование:</b> {name_column}\n{address}\n<b>Тип подключения:</b> {connection_type}\n<b>Скорость:</b> {speed}\n<b>Контакты:</b> {contacts}'
        thumb_url = get_thumb_url(functional_customer_column)
        espd_and_szoreg_results.append((relevance, 'espd', title, value_str, thumb_url))

    for row in found_szoreg_values:
        location_column = row[3]
        type_institution_column = row[7]
        full_address = row[5]
        title_column = row[8]
        location_ratio = fuzz.ratio(query.query.lower(), location_column.lower()) * 1.3 # Исправлено здесь
        type_institution_ratio = fuzz.ratio(query.query.lower(), type_institution_column.lower()) # Исправлено здесь
        title_ratio = fuzz.ratio(query.query.lower(), title_column.lower()) # Исправлено здесь
        max_ratio = max(location_ratio, type_institution_ratio, title_ratio)
        relevance = max_ratio
        type_institution = row[7]
        title = f"{type_institution} - {full_address}"
        address = row[5]
        connection_type = row[6]
        bandwidth = row[9]
        value_str = f'🏢<b>СЗО (региональный ГК)</b>\n\n<b>Тип учреждения:</b> {type_institution}\n<b>Наименование:</b> {title}\n<b>Адрес:</b> {full_address}\n<b>Тип подключения:</b> {connection_type}\n<b>Пропускная способность:</b> {bandwidth}'

        thumb_url = get_thumb_url(type_institution_column)
        espd_and_szoreg_results.append((relevance, 'szoreg', title, value_str, thumb_url))

    espd_and_szoreg_results.sort(reverse=True, key=lambda x: x[0])

    bot_url = "https://t.me/MCRcoffee_bot" # URL чата с ботом

    for _, _, title, value_str, thumb_url in espd_and_szoreg_results:
        result = types.InlineQueryResultArticle(
            id=str(len(results)),
            title=title,
            input_message_content=types.InputTextMessageContent(message_text=value_str, parse_mode='HTML'),
            thumb_url=thumb_url,
            reply_markup=types.InlineKeyboardMarkup().add(
                types.InlineKeyboardButton("Перейти в чат с ботом", url=bot_url) # Кнопка ведет в чат с ботом
            )
        )
        results.append(result)


    if not results:
        await query.answer(results=[], switch_pm_text="Результаты не найдены. Попробуйте другой запрос.", switch_pm_parameter="no_results")
        return

    results = results[:50]

    await query.answer(results)
'''
async def search_values(query):
    try:
        agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
        gc = await agcm.authorize()
        spreadsheet = await gc.open_by_key(SPREADSHEET_ID_1)
        # Открываем конкретный диапазон
        range_name = 'goroda2.0!A1:AQ1721'
        result = await spreadsheet.values_batch_get(range_name)
        rows = result.get('valueRanges', [])[0].get('values', [])
        normalized_query = normalize_text_v2(query)

        found_values_a = []
        found_values_s = []

        for row in rows:
            try:
                if len(row) > 0 and normalized_query == normalize_text_v2(row[0]):
                    found_values_a.append(row)
            except IndexError:
                pass

            try:
                if len(row) > 18 and fuzz.token_sort_ratio(normalized_query, normalize_text_v2(row[18])) >= 99:
                    found_values_s.append(row)
            except IndexError:
                pass

        return found_values_a, found_values_s
    except Exception as e:
        print("An error occurred during search_values:", e)
        return None, None


#test
async def search_values(query, spreadsheet):
    try:
        # Открываем конкретный диапазон
        range_name = 'goroda2.0!A1:AQ1721'
        result = await spreadsheet.values_batch_get(range_name)
        rows = result.get('valueRanges', [])[0].get('values', [])
        normalized_query = normalize_text_v2(query)

        found_values_a = []
        found_values_s = []

        for row in rows:
            try:
                if len(row) > 0 and normalized_query == normalize_text_v2(row[0]):
                    found_values_a.append(row)
            except IndexError:
                pass

            try:
                if len(row) > 18 and fuzz.token_sort_ratio(normalized_query, normalize_text_v2(row[18])) >= 99:
                    found_values_s.append(row)
            except IndexError:
                pass

        return found_values_a, found_values_s
    except Exception as e:
        print("An error occurred during search_values:", e)
        return None, None
'''


cached_values_a = None

async def load_values_a(spreadsheet):
    global cached_values_a
    try:
        if cached_values_a is None:
            range_name = 'goroda2.0!A1:AQ1721'
            result = await spreadsheet.values_batch_get(range_name)
            rows = result.get('valueRanges', [])[0].get('values', [])
            cached_values_a = rows
        return cached_values_a
    except Exception as e:
        print("An error occurred during loading values_a:", e)
        return None

async def search_values(query, spreadsheet):
    try:
        # Загрузить данные, если они еще не загружены
        values_a = await load_values_a(spreadsheet)

        # Нормализовать запрос для поиска
        normalized_query = normalize_text_v2(query.lower())

        # Найти нужные значения в загруженных данных
        found_values_a = [row for row in values_a if len(row) > 0 and normalized_query == normalize_text_v2(row[0].lower())]

        return found_values_a
    except Exception as e:
        print("An error occurred during search_values:", e)
        return None
'''
async def search_values(query, spreadsheet):
    try:
        # Сгенерировать ключ для кэширования на основе запроса
    #    cache_key = f"values_a:{query.lower()}"
     #   print(f"Cache key in search_values: {cache_key}")

        # Попытаться получить данные из кэша
     #   cached_data = await storage.get_data(chat=cache_key)
     #   if cached_data:
      #      cached_data = json.loads(cached_data)  # Десериализация данных
      #      print("Data found in cache")
      #      return cached_data

      #  print("Data not found in cache, fetching fresh data")

        # Если данных в кэше нет, продолжаем поиск в таблице
        range_name = 'goroda2.0!A1:AQ1721'
        result = await spreadsheet.values_batch_get(range_name)
        rows = result.get('valueRanges', [])[0].get('values', [])
        normalized_query = normalize_text_v2(query.lower())

        found_values_a = []


        for row in rows:
            try:
                if len(row) > 0 and normalized_query == normalize_text_v2(row[0].lower()):
                    found_values_a.append(row)
            except IndexError:
                pass

            try:
                if len(row) > 18 and fuzz.token_sort_ratio(normalized_query, normalize_text_v2(row[18].lower())) >= 99:
                    found_values_s.append(row)
            except IndexError:
                pass

        # Сохранить результаты в кэше перед возвратом
      #  await storage.set_data(chat=cache_key, data=json.dumps((found_values_a)))
     #   print("Data fetched and saved to cache")

        return found_values_a
    except Exception as e:
        print("An error occurred during search_values:", e)
        return None
'''





from fuzzywuzzy import fuzz

async def search_values_levenshtein(query, spreadsheet, threshold=0.7, max_results=5):
    try:
        # Открываем конкретный диапазон
        range_name = 'goroda2.0!A1:AM1721'
        result = await spreadsheet.values_batch_get(range_name)
        rows = result.get('valueRanges', [])[0].get('values', [])
        normalized_query = normalize_text_v2(query)

        # Создаем список для хранения всех совпадений
        all_matches = []

        for row in rows:
            try:
                if len(row) > 0:
                    similarity = fuzz.token_sort_ratio(normalized_query, normalize_text_v2(row[0]))
                    if similarity >= (threshold * 100):
                        all_matches.append((row, similarity))
            except IndexError:
                pass

        # Сортируем все совпадения по показателю сходства в обратном порядке (от большего к меньшему)
        sorted_matches = sorted(all_matches, key=lambda x: x[1], reverse=True)

        # Берем до max_results наиболее релевантных результатов
        top_matches = sorted_matches[:max_results]

        # Получаем только значения из первых позиций в каждом совпадении
        found_values_a = [match[0][0] for match in top_matches]

        return found_values_a
    except Exception as e:
        print("An error occurred during search_values_levenshtein:", e)
        return []




'''
@dp.message_handler()
async def search(message: types.Message):
    print ('async def search')
    query = message.text

    found_values_a, found_values_s = await search_values(query)

    if found_values_a:
        headers = ['Наименование', 'Население', 'Сотовая связь', 'Интернет', 'Программа', 'Таксофон', 'СЗО (узел)']
        data = [headers] + found_values_a
        excel_file = create_excel_file(headers, found_values_a)
        await message.answer_document(excel_file, caption='Результаты поиска (вариант А)')

    if found_values_s:
        headers = ['Наименование', 'Население', 'Сотовая связь', 'Интернет', 'Программа', 'Таксофон', 'СЗО (узел)']
        data = [headers] + found_values_s
        excel_file = convert_to_excel(data)
        await message.answer_document(excel_file, caption='Результаты поиска (вариант С)')
'''
#15
async def handle_mszu_message(message):
    print(f"Handling MSZU message: {message.text}")  #
    found_values = check_mszu_mo(message.text)
    if found_values:
        response = ""
        for i, row in enumerate(found_values):
            response += f"{i + 1}. Наименование ОМСУ из ЕСНСИ: {row[10]}\n" \
                        f"№ в Рег.перечне (17-р) : {row[0]}\n" \
                        f"Наименование услуги: {row[2]}\n" \
                        f"Ответ. РОИВ в рег.перечне: {row[3]}\n\n"

        # Разбиваем ответ на части, используя функцию split_message
        response_parts = split_message(response)

        # Отправляем каждую часть ответа по отдельности
        for part in response_parts:
            await bot.send_message(message.chat.id, part)
    else:
        await bot.send_message(message.chat.id, "Не удалось найти информацию. Попробуйте уточнить ваш запрос.")


async def handle_2mszu_message(message):
    print(f"Handling 2MSZU message: {message.text}")
    found_values = check_mszu_mo_2(message.text)
    if found_values:
        message_storage[message.chat.id] = message.text
        unique_values = list(set([row[12] for row in found_values]))
        unique_values.sort()

        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)

        numbered_list = ""
        for i, value in enumerate(unique_values):
            markup.add(types.KeyboardButton(str(i + 1)))  # Добавляем цифры вместо значений
            numbered_list += f"{i + 1}. {value}\n"

        msg = await bot.send_message(message.chat.id, f"Выберите необходимое учреждение:\n\n{numbered_list}", reply_markup=markup)
        await msg.answer_handler(handle_unique_value_selection)
    else:
        await bot.send_message(message.chat.id, "Не удалось найти информацию. Попробуйте уточнить ваш запрос.")


async def handle_unique_value_selection(message):
    user_message = message_storage[message.chat.id]
    selected_number = int(message.text)  # Получаем выбранный номер
    found_values = check_mszu_mo_2(user_message)
    unique_values = list(set([row[12] for row in found_values]))
    unique_values.sort()
    selected_value = unique_values[selected_number - 1]  # Получаем значение, соответствующее выбранному номеру

    selected_rows = [row for row in found_values if row[12] == selected_value]
    response = ""
    for i, row in enumerate(selected_rows):
        response += f"{i + 1}. Наименование ОМСУ из ЕСНСИ: {row[10]}\n" \
                    f"№ в Рег.перечне (17-р) : {row[0]}\n" \
                    f"Наименование услуги: {row[2]}\n" \
                    f"Ответ. РОИВ в рег.перечне: {row[3]}\n\n"

    response_parts = split_message(response)

    for part in response_parts:
        if part.strip():
            await bot.send_message(message.chat.id, part)


user_messages = {}

'''
async def handle_tor_message(message):
    query = message.text[4:]  # Получаем часть сообщения после "тор "
    service = build('sheets', 'v4', credentials=creds)

    # Получаем заголовок таблицы
    header_result = await service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID_1, range='nst!H1:X1'
    ).execute()
    column_headers = header_result.get('values', [])[0]

    # Поиск совпадений в столбце A на листе nst
    nst_result = await service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID_1, range='nst!A1:X150'
    ).execute()
    nst_rows = nst_result.get('values', [])
    nst_matches = [
        row for row in nst_rows if row and normalize_text_v2(query) == normalize_text_v2(row[0])
    ]

    if nst_matches:
        all_responses = []
        total_yes_count = 0
        total_no_count = 0

        for row in nst_matches:
            institution_name = row[5]
            response_parts = [f"{institution_name}:\n\n"]
            yes_count = 0
            no_count = 0

            for i in range(7, 24):
                column_header = column_headers[i - 7]
                cell_value = row[i]

                if cell_value == "Есть":
                    cell_value = f"🟢{column_header}\n"
                    yes_count += 1
                elif cell_value == "Нет":
                    cell_value = f"🔴{column_header}\n"
                    no_count += 1

                response_parts.append(cell_value)

            response = " ".join(response_parts)
            all_responses.append(response)
            total_yes_count += yes_count
            total_no_count += no_count

        # Create and save the pie chart for the total counts
        pie_chart_filename = "pie_chart_total.png"
        create_pie_chart(total_yes_count, total_no_count, pie_chart_filename)

        # Combine all responses into one message
        combined_response = "\n\n".join(all_responses)
        messages = split_message(combined_response)

        # Send the pie chart
        await bot.send_message(message.chat.id, 'Диаграмма подключения к ТОРКНД')
        # Создаем inline кнопку "подробно"
        markup = types.InlineKeyboardMarkup()
        detailed_button = types.InlineKeyboardButton("подробно", callback_data='{"type": "additional_info_tor"}')
        markup.add(detailed_button)

        # Отправляем изображение с кнопкой
        with open(pie_chart_filename, 'rb') as chart_file:
            callback_data = json.dumps({
                "type": "additional_info_tor",
                "user_id": message.from_user.id
            })
            detailed_button = types.InlineKeyboardButton("подробно", callback_data=callback_data)
            markup = types.InlineKeyboardMarkup()
            markup.add(detailed_button)

            # Сохраняем сообщения для пользователя в глобальном словаре
            user_messages[message.from_user.id] = messages

            await bot.send_photo(message.chat.id, chart_file, reply_markup=markup)

        # Remove the pie chart file after sending
        os.remove(pie_chart_filename)
        user_messages[message.chat.id] = messages
    else:
        await bot.send_message(message.chat.id, "Не найдено совпадений для запроса")

#16

async def handle_mszu_choice(message, found_mszu_values, keyboard):
    print('async def handle_mszu_choice(message, found_mszu_values, keyboard):')
    choice = message.text.strip()

    if choice.isdigit():
        index = int(choice) - 1
        if 0 <= index < len(found_mszu_values):
            response = f"\nНаименование услуги: \n{found_mszu_values[index][2]}\n№ в Рег.перечне (17-р): {found_mszu_values[index][0]}\nОтвет. РОИВ в рег.перечне: {found_mszu_values[index][3]}\nОтв. в плане: {found_mszu_values[index][6]}\nЕСНСИ: {found_mszu_values[index][8]}\nЕПГУ: {found_mszu_values[index][10]}"
            await bot.send_message(message.chat.id, response, reply_markup=types.ReplyKeyboardRemove())
            inline_keyboard = types.InlineKeyboardMarkup()
            url_button = types.InlineKeyboardButton(text="Адрес ИФЗ", url=found_mszu_values[index][13])
            inline_keyboard.add(url_button)

            # Отправка сообщения с inline кнопкой
            await bot.send_message(message.chat.id, "Нажмите на кнопку ниже для перехода по ссылке:", reply_markup=inline_keyboard)
        else:
            await bot.send_message(message.chat.id, 'Неверный номер. Попробуйте еще раз или нажмите "Отмена".', reply_markup=keyboard)
            await bot.register_next_step_handler(message, lambda x: handle_mszu_choice(x, found_mszu_values, keyboard))
    elif choice.lower() == "отмена":
        await bot.send_message(message.chat.id, 'Поиск отменен.', reply_markup=types.ReplyKeyboardRemove())
    else:
        await bot.send_message(message.chat.id, 'Введите корректный номер или нажмите "Отмена".', reply_markup=keyboard)
        await bot.register_next_step_handler(message, lambda x: handle_mszu_choice(x, found_mszu_values, keyboard))


@dp.callback_query_handler(lambda call: json.loads(call.data)["type"] == "additional_info_tor")
async def detailed_button_callback(call):
    user_first_name = call.from_user.first_name
    await bot.send_message(call.message.chat.id, f'Секундочку, {user_first_name}😌 Загружаю статистику для тебя')
    time.sleep(3)

    # Получаем messages из глобального словаря
    callback_data = json.loads(call.data)
    user_id = callback_data["user_id"]
    messages = user_messages.get(user_id, [])

    for msg in messages:
        await bot.send_message(call.message.chat.id, msg)
    await bot.answer_callback_query(call.id)
    time.sleep(2)
    await bot.send_message(call.message.chat.id, 'Введите свой следующий запрос')


async def handle_knd_kraj_message(message):
    service = build('sheets', 'v4', credentials=creds)

    # Получаем все строки таблицы nst
    nst_result = await service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID_1, range='nst!A1:X150'
    ).execute()
    nst_rows = nst_result.get('values', [])

    # Подсчет количества значений "Есть" и "Нет" для каждой строки в столбце A
    counter = {}
    for row in nst_rows:
        key = row[0]
        yes_count = sum([1 for value in row[7:24] if value == "Есть"])
        no_count = sum([1 for value in row[7:24] if value == "Нет"])

        if key in counter:
            counter[key] = (key, yes_count + counter[key][1], no_count + counter[key][2])
        else:
            counter[key] = (key, yes_count, no_count)

    # Создание столбчатой диаграммы
    bar_chart_filename = 'bar_chart.png'
    create_bar_chart(list(counter.values()), bar_chart_filename)

    # Отправка столбчатой диаграммы
    with open(bar_chart_filename, 'rb') as chart_file:
        await bot.send_photo(message.chat.id, chart_file)

    # Удаление файла диаграммы после отправки
    os.remove(bar_chart_filename)
'''

COLUMNS_TO_EXPORT = [1, 2, 3, 6, 7]

# Заголовки таблицы
TABLE_HEADERS = ["Наименование", "Население", "Сотовая связь", "Интернет", "Таксофон"]


async def handler_fp_message(message, fp_data, fp_headers):
    time.sleep(1)
    try:
        data = [
            [row[0], row[1], row[2], row[4], row[5]]
            for row in fp_data[1:]
        ]
        headers = ['Муниципальное образование', 'Наименование населенного пункта', 'Адрес', 'Наименование учреждения', 'Скорость']

        excel_file = create_excel_file(headers, data)
        user_first_name = message.from_user.first_name
        await bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌')
        time.sleep(2)
        await bot.send_message(message.chat.id, 'Перечень ФАП из госконтракта:')
        file_name = 'ФАП'

        with BytesIO(excel_file) as file:
            file.name = f'{file_name}.xlsx'
            await bot.send_document(message.chat.id, file)

    except Exception as e:
        logging.exception(e)
        await bot.send_message(message.chat.id, "Произошла ошибка. Попробуйте еще раз.")


async def handler_aggoics_message(message, aggoics_data, aggoics_headers):
    time.sleep(1)
    try:
        data = [
            [row[0], row[1], row[2], row[3], row[4], row[5]]
            for row in aggoics_data[1:]
        ]

        excel_file = create_excel_file(aggoics_headers, data)

        await bot.send_message(message.chat.id, 'Информация из таблицы АгГОиЧС:')
        file_name = 'АгГОиЧС'

        with BytesIO(excel_file) as file:
            file.name = f'{file_name}.xlsx'
            await bot.send_document(message.chat.id, file)

    except Exception as e:
        logging.exception(e)
        await bot.send_message(message.chat.id, "Произошла ошибка. Попробуйте еще раз.")



#17

async def handler_otpusk_message(message, employees_on_vacation):
    if len(employees_on_vacation) > 0:
        response = "Сотрудники, которые сегодня в отпуске:\n\n"
        for employee in employees_on_vacation:
            response += f"{employee[0]} ({employee[1]})\n"
        time.sleep(2)
        await message.reply(response)
    else:
        time.sleep(2)
        await message.reply("Сегодня никто не в отпуске.")





@dp.message_handler(state=Form.waiting_for_number)
async def handle_choice(message: types.Message, state: FSMContext):
    start_time = time.time()
    try:
        data = await state.get_data()
        found_values = data.get('found_values')

        index_text = message.text
        user_first_name = message.from_user.first_name
        chat_id = message.chat.id
        response = ""
        pokazatel_504p_lines = []
        votes_response = ""
        yandex_2023_response = ""
        ucn2_response = ""

        # Проверки ввода пользователя
        if index_text == "Отмена":
            await bot.send_message(chat_id, 'Поиск отменен.', reply_markup=types.ReplyKeyboardRemove())
            await state.reset_state()
            return
        if not index_text.isdigit():
            await bot.send_message(chat_id, 'Введено некорректное значение. Пожалуйста, введите число.')
            return

        index = int(index_text)
        if index <= 0 or index > len(found_values):
            await bot.send_message(chat_id, f'Введено некорректное значение. Пожалуйста, введите число в диапазоне от 1 до {len(found_values)}.')
            return

        selected_np = found_values[index - 1]
        latitude = selected_np[7]
        longitude = selected_np[8]

        # Получаем все необходимые данные асинхронно
        gc, spreadsheet = await get_authorized_client_and_spreadsheet()

        weather_info, yandex_2023_values, pokazatel_504p_values, survey_results_values, ucn2_values = await asyncio.gather(
            get_weather(latitude, longitude, "7cc8daec601b8354e0bc6350592d6f98"),
            search_yandex_2023_values(selected_np[4], spreadsheet),
            search_in_pokazatel_504p(selected_np[4], spreadsheet),
            search_in_results(selected_np[4], spreadsheet),

            search_in_ucn2(selected_np[4], spreadsheet)
        )










        # Создаем словарь с операторами и их значениями
        operators = {
            "Tele2": selected_np[39] if len(selected_np) > 39 else None,
            "Мегафон": selected_np[40] if len(selected_np) > 40 else None,
            "Билайн": selected_np[41] if len(selected_np) > 41 else None,
            "МТС": selected_np[42] if len(selected_np) > 42 else None,
        }

        operators_response = '\nОценка жителей:\n'

        # Список для хранения ответов от операторов
        operator_responses = []

        for operator_name, operator_value in operators.items():

            if operator_value:  # Проверка на наличие значения (не None и не пустая строка)
                # Переводим значение в строку, чтобы избежать ошибок при выполнении метода replace
                operator_value_str = str(operator_value)

                # Пытаемся найти качество сигнала
                signal_quality = re.search(r'Отсутствует|Низкое|Среднее|Хорошее', operator_value_str, re.IGNORECASE)
                if signal_quality:
                    signal_quality = signal_quality.group()
                    if "Отсутствует" in signal_quality:
                        signal_level = "🔴Отсутствует"
                    elif "Низкое" in signal_quality:
                        signal_level = "🟠Низкое"
                    elif "Среднее" in signal_quality:
                        signal_level = "🟡Среднее"
                    elif "Хорошее" in signal_quality:
                        signal_level = "🟢Хорошее"
                    else:
                        signal_level = "❓Неизвестно"

                    # Заменяем найденное качество сигнала на его эмодзи-эквивалент
                    operator_value_str = operator_value_str.replace(signal_quality, signal_level)
                else:
                    operator_value_str = operator_value_str

                # Заменяем "(" и ")" на " "
                operator_value_str = operator_value_str.replace("(", " ").replace(")", " ")
                operator_responses.append(f'{operator_name}: {operator_value_str}\n')
            else:
                # Если нет данных по оператору, не добавляем его в ответ
                continue



        # Если нет данных ни по одному оператору, добавляем сообщение об отсутствии данных
        if not operator_responses:
            operators_response += 'Нет данных\n'
        else:
            operators_response += ''.join(operator_responses)

        response += operators_response






        if yandex_2023_values:
            yandex_2023_response = '\n\n<b>Информация из таблицы 2023</b>\n\n'
            for row in yandex_2023_values:
                yandex_2023_response += f'Тип подключения: {row[4]}\nОператор: {row[15]}\nСоглашение: {row[7]}\nПодписание соглашения с МЦР: {row[8]}\nПодписание соглашения с АГЗ: {row[9]}\nДата подписания контракта: {row[11]}\nДата установки АМС: {row[12]}\nДата монтажа БС: {row[13]}\nЗапуск услуг: {row[14]}\n\n'

        if len(pokazatel_504p_values) > 0:
            for i in range(6, 10):
                if len(pokazatel_504p_values[0]) > i and pokazatel_504p_values[0][i] and pokazatel_504p_values[0][i].strip():
                    value = pokazatel_504p_values[0][i]
                    value = value.replace("Хорошее", "🟢Хорошее").replace("Низкое", "🟠Низкое")
                    pokazatel_504p_lines.append(f"{value}")

        pokazatel_504p_response = "\n".join(pokazatel_504p_lines) if pokazatel_504p_lines else "🔴отсутствует"



        if "4G" in pokazatel_504p_response:
            votes_response = ""
        else:
            if len(selected_np) > 38:
                votes = selected_np[34] or "неизвестно"  # Количество голосов находится в 35-ом столбце
                update_time = selected_np[35] or "неизвестно"  # Время обновления находится в 36-ом столбце
                rank = selected_np[36] or "неизвестно"  # Рейтинг находится в 37-ом столбце
                same_votes_np = selected_np[38] or "неизвестно"  # Количество НП с таким же количеством голосов находится в 39-ом столбце
                if votes != "неизвестно" and update_time != "неизвестно" and rank != "неизвестно" and same_votes_np != "неизвестно":
                    votes_response = f'\n\n<b>Голосование УЦН 2.0 2024</b>\n\n📊Количество голосов: <b>{votes}</b> (такое же количество голосов имеют {same_votes_np} населённых пунктов)\n🏆Место в рейтинге: {rank}\nДата обновления информации: {update_time}'
                else:
                    print("Debug: Не все данные для блока голосования были найдены.")



        response += votes_response


        '''
        if len(selected_np) > 38:
            votes = selected_np[34] or "неизвестно"  # Количество голосов находится в 35-ом столбце
            update_time = selected_np[35] or "неизвестно"  # Время обновления находится в 36-ом столбце
            rank = selected_np[36] or "неизвестно"  # Рейтинг находится в 37-ом столбце
            same_votes_np = selected_np[38] or "неизвестно"  # Количество НП с таким же количеством голосов находится в 39-ом столбце
            if votes != "неизвестно" and update_time != "неизвестно" and rank != "неизвестно" and same_votes_np != "неизвестно":
                votes_response = f'\n\n<b>Голосование УЦН 2.0 2024</b>\n\n📊Количество голосов: <b>{votes}</b> (такое же количество голосов имеют {same_votes_np} населённых пунктов)\n🏆Место в рейтинге: {rank}\nДата обновления информации: {update_time}'

            else:
                print("Не все данные для блока голосования были найдены.")
        response += votes_response
        '''



        if ucn2_values:
            for row in ucn2_values:
                ucn2_response = ''

                if 4 < len(row) and row[4]:  # Проверка наличия значения
                    ucn2_response += '  Информация от Теле2:\n    -СМР: ' + row[4] + '\n'
                if 5 < len(row) and row[5]:  # Проверка наличия значения
                    ucn2_response += '    -Запуск: ' + row[5] + '\n'
                if 6 < len(row) and row[6]:  # Проверка наличия значения
                    ucn2_response += '    -Комментарии: ' + row[6] + '\n'

                if ucn2_response:  # Если ucn2_response не пуст, добавить вводную строку в начало
                    ucn2_response = '\n\n\n<b>УЦН 2.0 2023</b>\n' + ucn2_response
                    response += ucn2_response



        survey_data_storage[message.chat.id] = survey_results_values


        try:
            selsovet_info, tanya_sub_info_year, tanya_sub_info_provider, taksofony_info, arctic_info, internet_info, population_2010, population_2020, itog_ucn_2023 = await asyncio.gather(
                get_value(found_values[index - 1], 20),
                get_value(found_values[index - 1], 13),
                get_value(found_values[index - 1], 14),
                get_value(found_values[index - 1], 12),
                get_value(found_values[index - 1], 6),
                get_value(found_values[index - 1], 9),
                get_value(found_values[index - 1], 2),
                get_value(found_values[index - 1], 5),
                get_value(found_values[index - 1], 24),
                return_exceptions=True  # Возврат исключений как объектов
            )
        except Exception as e:
            print(f"Произошла ошибка: {e}")

     #   response = f'<b>{await get_value(found_values[index - 1], 1)}</b> {weather_info}\n\n👥Население (2010 г): {await get_value(found_values[index - 1], 2)} чел.\n👥Население(2020 г): {await get_value(found_values[index - 1], 5)} чел.\n\n<b>Сотовая связь:</b>\n{pokazatel_504p_response}\n{operators_response}\n\nИнтернет: {await get_value(found_values[index - 1], 9)}\n\nКоличество таксофонов: {await get_value(found_values[index - 1], 12)}{ucn2_response}{yandex_2023_response}{votes_response}\n\nЕсли хочешь узнать о голосовании УЦН 2.0 2024 жми /votes\nБот для проведения опросов жителей - <a href="http://t.me/providers_rating_bot">@providers_rating_bot</a>'

        response = f'<b>{await get_value(found_values[index - 1], 1)}</b>  {weather_info}'

        if selsovet_info:
            response += f'\n{selsovet_info}'


        if arctic_info:
            response += f'\n❄️️арктическая зона❄️️'


        response += f'\n\n👥население 2010 г: {population_2010} чел.\n👥население 2020 г: {population_2020} чел.'

        if taksofony_info:
                response += f'\n☎️таксофон: {taksofony_info}'

        response += f'\n🌐интернет: {internet_info}️'
        response += f'\n\n📱<b>Сотовая связь:</b>\n{pokazatel_504p_response}'


        if tanya_sub_info_year and tanya_sub_info_provider:
            response += f'\n\nнаселенный пункт был подключен в рамках государственной программый "Развитие информационного общества" в {tanya_sub_info_year} году, оператор {tanya_sub_info_provider}'

        if itog_ucn_2023:
            response += f'\n\nкомментарий Минцифры России об УЦН 2024: {itog_ucn_2023}'

        response += f'\n{operators_response}\n'

        response += f'{ucn2_response}{yandex_2023_response}{votes_response}\nЕсли хочешь узнать о голосовании УЦН 2.0 2024 жми /votes\nБот для проведения опросов жителей - <a href="http://t.me/providers_rating_bot">@providers_rating_bot</a>'



        info_text_storage[message.chat.id] = response






        await bot.send_message(message.chat.id, "<b>Выбранный населенный пункт</b>", parse_mode='HTML', reply_markup=types.ReplyKeyboardRemove())
        await bot.send_location(message.chat.id, latitude, longitude)
        #response = await send_request_to_gpt(chat_id, response)
        messages = split_message(response)


        allowed_users = {964635576, 1063749463, 374056328, 572346758, 434872315, 1045874687, 1063749463, 487922464, 371098269, 402748716}
        for msg in messages:
            await bot.send_message(message.chat.id, msg, parse_mode='HTML', disable_web_page_preview=True)




        '''
        szofed_values, espd_values, szoreg_values, schools_values = await asyncio.gather(

            search_szofed_values(selected_np[4], spreadsheet),
            search_espd_values(selected_np[4], spreadsheet),
            search_szoreg_values(selected_np[4], spreadsheet),
            search_schools_values(selected_np[4], spreadsheet)

        )
        '''
        szoreg_values, schools_values = await asyncio.gather(


            search_szoreg_values(selected_np[4], spreadsheet),
            search_schools_values(selected_np[4], spreadsheet)

        )





        await state.reset_state()
        inline_keyboard = types.InlineKeyboardMarkup()

        if survey_results_values:
            survey_inline_keyboard = types.InlineKeyboardMarkup()
            button_survey_results = types.InlineKeyboardButton("Показать результаты опроса", callback_data=json.dumps({"type": "survey_chart", "chat_id": message.chat.id}))
            survey_inline_keyboard.add(button_survey_results)
            await bot.send_message(message.chat.id, "Найдены результаты опроса. Хотите посмотреть?", reply_markup=survey_inline_keyboard)

        if message.from_user.id in allowed_users:
            button_digital_ministry_info = types.InlineKeyboardButton("😈Подготовить ответ на обращение(БЕТА)", callback_data=json.dumps({"type": "digital_ministry_info", "chat_id": message.chat.id}))
            inline_keyboard.add(button_digital_ministry_info)
          #  button_digital_ministry_info_post = types.InlineKeyboardButton("Сделать пост ВК", callback_data=json.dumps({"type": "digital_ministry_info_post", "chat_id": message.chat.id}))
           # inline_keyboard.add(button_digital_ministry_info_post)
        '''
        if szofed_values:
            szofed_response = '*🏢СЗО, подключенные в рамках федерального проекта в период с 2019 по 2021 год:*\n\n'
            for i, row in enumerate(szofed_values, 1):
                szofed_response += f'\n{i}. *Тип:* {row[8]}\n*Наименование:* {row[9]}\n*Адрес:* {row[4]}\n*Тип подключения (Узел связи):* {row[10]}\n*Пропускная способность:* {row[11]} Мб/сек\n*Дата подключения:* {row[12]}\n'

            callback_data = json.dumps({"type": "additional_info", "chat_id": message.chat.id})
            additional_info_storage[message.chat.id] = szofed_response
            button_additional_info = types.InlineKeyboardButton("🏢СЗО", callback_data=callback_data)
            inline_keyboard.add(button_additional_info)

        if espd_values:
            espd_response = '🌐<b>Точки подключения к ЕСПД:</b> \n\n'
            for i, row in enumerate(espd_values, 1):
                espd_response += f'\n{i}. <b>Наименование:</b> {html.escape(row[12])}\n<b>Адрес:</b> {html.escape(row[8])}\n<b>Тип подключения:</b> {html.escape(row[9])}\n<b>Скорость:</b> {html.escape(row[13])}\n<b>Контакты:</b> {html.escape(row[18])}\n'

            callback_data = json.dumps({"type": "espd_info", "chat_id": message.chat.id})
            espd_info_storage[message.chat.id] = espd_response
            button_espd_info = types.InlineKeyboardButton("🌐ЕСПД", callback_data=callback_data)
            inline_keyboard.add(button_espd_info)
        '''

        if szoreg_values:
            szoreg_response = '🏢<b>Учреждения, подключенные по госпрограмме</b>\n\n'
            for i, row in enumerate(szoreg_values, 1):

                szoreg_response += f'\n{i}. <b>Тип:</b> {row[7]}\n<b>аименование:</b> {row[8]}\n<b>Адрес:</b> {row[5]} \n<b>Тип подключения:</b> {row[6]}\n<b>Пропускная способность:</b> {row[9]}\n<b>Контракт:</b> {row[10]}\n'

               # if len(row) >= 11:

                 #   szoreg_response += f'<b>Комментарий:</b> {row[11]}\n'


            callback_data = json.dumps({"type": "szoreg_info", "chat_id":message.chat.id})
            szoreg_info_storage[message.chat.id] = szoreg_response
            button_szoreg_info = types.InlineKeyboardButton(f"🏢Список учреждений ({len(szoreg_values)})", callback_data=callback_data)
            inline_keyboard.add(button_szoreg_info)

        if schools_values:
            schools_response = '🏫<b>Школы:</b>\n'
            for i, row in enumerate(schools_values, 1):
                schools_response += f'\n\n{i}. '
                if len(row) > 7:
                    schools_response += f'<b>{html.escape(row[12])}</b>\n'
                if len(row) > 12:
                    schools_response += f'\n{html.escape(row[7])}\n'
                if len(row) > 14:
                    schools_response += f'\n{html.escape(row[14])}, '
                if len(row) > 13:
                    schools_response += f'{html.escape(row[13])} Мб/с\n'
                if len(row) > 20:
                    schools_response += f'{html.escape(row[20])}'
                schools_response += '\n'

            callback_data = json.dumps({"type": "schools_info", "chat_id": message.chat.id})
            schools_info_storage[message.chat.id] = schools_response
            button_schools_info = types.InlineKeyboardButton("🏫Школы",callback_data=callback_data)
            inline_keyboard.add(button_schools_info)

        await bot.send_message(message.chat.id, "⬇️Дополнительная информация⬇️", reply_markup=inline_keyboard)


      #  dp.register_callback_query_handler(handle_additional_info, lambda query: json.loads(query.data)["type"] == "additional_info")
       # dp.register_callback_query_handler(handle_espd_info, lambda query: json.loads(query.data)["type"] == "espd_info")
    #    dp.register_callback_query_handler(handle_szoreg_info, lambda query: json.loads(query.data)["type"] == "szoreg_info")
     #   dp.register_callback_query_handler(handle_digital_ministry_info, lambda query: json.loads(query.data)["type"] == "digital_ministry_info")
      #  dp.register_callback_query_handler(handle_digital_ministry_info_post, lambda query: json.loads(query.data)["type"] == "digital_ministry_info_post")
       # dp.register_callback_query_handler(handle_schools_info, lambda query: json.loads(query.data)["type"] == "schools_info")
    except ValueError:
        # Если пользователь ввел число, которое не входит в ожидаемый диапазон
        await bot.send_message(message.chat.id, 'Введено некорректное значение. Пожалуйста, введите число в диапазоне от 1 до {}.'.format(len(found_values)))




dp.register_callback_query_handler(handle_survey_chart, lambda query: json.loads(query.data)["type"] == "survey_chart")

from aiogram import executor
async def on_startup(dp):
    #await clear_cache()
    print("Бот запущен")
'''
async def clear_cache():
    # Добавьте asyncio.loop.run_in_executor для работы с синхронным кодом в асинхронной функции
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, storage2.flushdb)
    print("Cache cleared")
'''

if __name__ == "__main__":
    # Настройка и запуск планировщика
   # scheduler = AsyncIOScheduler()
    #scheduler.add_job(clear_cache, 'cron', hour=0, minute=0)
    #scheduler.start()

    executor.start_polling(dp, on_startup=on_startup)




'''
async def set_webhook():
    await bot.set_webhook(WEBHOOK_URL, certificate=open(WEBHOOK_SSL_CERT, 'r'))
    print("Webhook set successfully")




async def main():
    try:
        await set_webhook()

        print(bot.get_webhook_info)
        print("Webhook has been set.")


        # Для локальной отладки
        # await dp.skip_updates()
        # await dp.start_polling()

    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    asyncio.run(main())






async def on_startup(dispatcher):
    try:
        await bot.set_webhook(WEBHOOK_URL, certificate=open(WEBHOOK_SSL_CERT, 'rb'))
        print("Webhook set up successfully.")
    except Exception as e:
        print(f"Error setting up webhook: {e}")
    # Другие действия при запуске

async def on_shutdown(dispatcher):
    try:
        await bot.delete_webhook()
        print("Webhook deleted successfully.")
    except Exception as e:
        print(f"Error deleting webhook: {e}")

async def clear_cache():
    try:
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, storage2.flushdb)
        print("Cache cleared")
    except Exception as e:
        print(f"Error clearing cache: {e}")

if __name__ == '__main__':
    try:
        scheduler = AsyncIOScheduler()
        scheduler.add_job(clear_cache, 'cron', hour=0, minute=0)
        scheduler.start()
        print("Scheduler started.")

        start_webhook(
            dispatcher=dp,
            webhook_path=WEBHOOK_PATH,
            on_startup=on_startup,
            on_shutdown=on_shutdown,
            host=WEBAPP_HOST,
            port=WEBAPP_PORT,
        )
    except Exception as e:
        print(f"Error in main: {e}")


'''