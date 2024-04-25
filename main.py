import traceback

from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from google.oauth2 import service_account
from aiogram import executor
import openpyxl
from openpyxl.utils import get_column_letter
from aiogram import executor
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from google.oauth2 import service_account
import gspread_asyncio

from aiogram.contrib.fsm_storage.memory import MemoryStorage
from google_connections import init_redis, load_szoreg_values, load_yandex_2023_values, load_pokazatel_504p_values, load_ucn2_values, load_schools_values, load_votes_values, load_survey_values, load_values, load_otpusk_data, SPREADSHEET_ID


from aiogram import types

from google.oauth2 import service_account
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from aiogram.types import InputFile

from config import bot_token

response_storage = {}
bot = Bot(token=bot_token)
dp = Dispatcher(bot, storage=MemoryStorage())
info_text_storage = {}




@dp.message_handler(commands=['help'])
async def handle_help_command(message: types.Message):
    await log_user_data_from_message(message)
    help_text = (
        'Введи название населенного пункта или муниципального образования, чтобы получить информацию о связи'
        'Чтобы узнать, кто сегодня в отпуске, жми /otpusk\n\n'
        'Если остались вопросы, пиши @rejoller.')
    await message.reply(help_text)





SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
SERVICE_ACCOUNT_FILE = 'credentials.json'

creds = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)


headers = ['Наименование', 'Население', 'Сотовая связь', 'Интернет', 'Программа', 'Таксофон', 'СЗО (узел)']


from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter







def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            value = str(cell.value)
            length = len(value.encode('utf-8'))
            if length > max_length:
                max_length = length

        # Настройка ширины столбца
        estimated_width = max_length * 0.7  # Умножение на коэффициент для учета разных ширин символов
        worksheet.column_dimensions[column].width = estimated_width


def convert_to_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    # Настраиваем стили для заголовков
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="AED6F1",
                              end_color="AED6F1",
                              fill_type="solid")

    # Настраиваем стили для данных
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    data_fill = PatternFill(start_color="ECECEC",
                            end_color="ECECEC",
                            fill_type="solid")

    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Применяем стили
            if row_idx == 1:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                cell.fill = header_fill
            else:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                if row_idx % 2 == 0:
                    cell.fill = data_fill

    # Вызов функции для автоматической настройки ширины столбцов
    adjust_column_width(ws)

    # Добавляем автофильтр
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Закрепляем строку заголовка
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@dp.message_handler(commands=['otpusk'])
async def handle_otpusk_command(message: types.Message):
    await message.answer('Загружаю данные')
    otpusk_data = await load_otpusk_data()
    await filter_and_send_data(message, otpusk_data)


async def filter_and_send_data(message: types.Message, data, headers=None):
    if headers:
        filtered_data = [headers] + data
    else:
        filtered_data = data

    filtered_data = [list(map(str, row)) for row in filtered_data]

    # Отправляем данные частями, чтобы не превысить лимит сообщения
    for i in range(0, len(filtered_data), 10):
        chunk = filtered_data[i:i+10]
        text = '\n'.join(['\t'.join(row) for row in chunk])
        await message.answer(text, parse_mode=types.ParseMode.MARKDOWN)


@dp.message_handler(commands=['employees_vacation'])
async def handle_employees_vacation_command(message: types.Message):
    await message.answer('Загружаю данные')
    otpusk_data = await load_otpusk_data()
    employees_on_vacation, employees_starting_vacation_soon = get_employees_on_vacation(otpusk_data)

    if employees_on_vacation:
        await message.answer('Сотрудники, находящиеся в отпуске:')
        await message.answer('\n'.join(['\t'.join(row) for row in employees_on_vacation]), parse_mode=types.ParseMode.MARKDOWN)
    else:
        await message.answer('Сотрудников в отпуске нет.')

    if employees_starting_vacation_soon:
        await message.answer('Сотрудники, начинающие отпуск в ближайшие дни:')
        await message.answer('\n'.join(['\t'.join(row) for row in employees_starting_vacation_soon]), parse_mode=types.ParseMode.MARKDOWN)
    else:
        await message.answer('Сотрудников, начинающих отпуск в ближайшие дни, нет.')




def escape_markdown(text):
    markdown_escape_characters = ['*', '_', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    return re.sub('([{}])'.format(''.join(markdown_escape_characters)), r'\\\1', text)


is_main_menu_button_active = {}







#13
from handlers import handle_additional_info, handle_szoreg_info, handle_schools_info, handle_survey_chart
dp.register_callback_query_handler(handle_additional_info, lambda query: json.loads(query.data)["type"] == "additional_info")
dp.register_callback_query_handler(handle_szoreg_info, lambda query: json.loads(query.data)["type"] == "szoreg_info")
dp.register_callback_query_handler(handle_schools_info, lambda query: json.loads(query.data)["type"] == "schools_info")
dp.register_callback_query_handler(handle_survey_chart, lambda query: json.loads(query.data)["type"] == "survey_chart")







user_messages = {}




TABLE_HEADERS = ["Наименование", "Население", "Сотовая связь", "Интернет", "Таксофон"]


async def handler_otpusk_message(message, employees_on_vacation):
    if len(employees_on_vacation) > 0:
        response = "Сотрудники, которые сегодня в отпуске:\n\n"
        for employee in employees_on_vacation:
            response += f"{employee[0]} ({employee[1]})\n"
        time.sleep(2)
        await message.reply(response)
    else:
        time.sleep(2)
        await message.reply("Сегодня никто не в отпуске.")











async def on_startup(dp):
    try:
        print('Initializing Redis...')
        redis = await init_redis()
        agcm = gspread_asyncio.AsyncioGspreadClientManager(lambda: creds)
        gc = await agcm.authorize()
        spreadsheet = await gc.open_by_key(SPREADSHEET_ID)
        await load_values(spreadsheet, redis)
        await load_szoreg_values(spreadsheet, redis)
        await load_pokazatel_504p_values(spreadsheet, redis)
        await load_schools_values(spreadsheet, redis)
        await load_yandex_2023_values(spreadsheet, redis)
        await load_ucn2_values(spreadsheet, redis)
        await load_survey_values(spreadsheet, redis)
        await load_votes_values(spreadsheet, redis)
        
        print('Initialization and data loading complete.')
        
    except Exception as e:
        print('Failed to initialize and load data:', str(e))
        traceback.print_exc()



if __name__ == "__main__":
    from handlers import *
    executor.start_polling(dp, on_startup=on_startup)
    



